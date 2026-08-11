#!/usr/bin/env python3
"""
Konfigurierbarer Generator für ETL-, Event-Log- und Process-Mining-Testdaten.

Erzeugt eine Excel-Datei mit den Tabellenblättern:
- Ereignisdaten
- Ressourcen
- Hinweise

Voraussetzung:
    pip install openpyxl

Beispiel:
    python testdatensatz_generator.py --faelle 2000 --ausgabe ETL_Testdaten_2000_Faelle.xlsx
"""

from __future__ import annotations

import argparse
import random
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


@dataclass(frozen=True)
class GeneratorKonfiguration:
    faelle: int
    ausgabe: Path
    seed: int
    startzeitpunkt: datetime
    anteil_standard: float = 0.60
    anteil_kurz: float = 0.20
    anteil_nacharbeit: float = 0.15
    anteil_transport_loop: float = 0.05
    fehlende_bearbeitungszeiten: int = 0
    ausreisser: int = 0
    platzhalter_ressource: int = 0
    platzhalter_status: int = 0
    fehlende_aktivitaeten: int = 0
    fehlende_zeitstempel: int = 0
    fehlende_lagerplaetze: int = 0
    platzhalter_lagerplaetze: int = 0
    identische_zeitstempel_faelle: int = 0
    ruecklaeufige_zeit_faelle: int = 0
    exakte_duplikate: int = 0


def argumente_lesen() -> GeneratorKonfiguration:
    parser = argparse.ArgumentParser(
        description="Erzeugt einen großen Excel-Testdatensatz für ETL und Process Mining."
    )
    parser.add_argument(
        "--faelle",
        type=int,
        default=2000,
        help="Anzahl zu erzeugender Aufträge/Fälle (Standard: 2000).",
    )
    parser.add_argument(
        "--ausgabe",
        type=Path,
        default=Path("ETL_Testdaten_generiert.xlsx"),
        help="Pfad der Ausgabedatei.",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=20260724,
        help="Zufallsseed für reproduzierbare Daten.",
    )
    parser.add_argument(
        "--start",
        type=str,
        default="2026-07-01 06:00:00",
        help="Startzeitpunkt im Format YYYY-MM-DD HH:MM:SS.",
    )
    args = parser.parse_args()

    if args.faelle < 1:
        parser.error("--faelle muss mindestens 1 sein.")

    try:
        startzeitpunkt = datetime.strptime(args.start, "%Y-%m-%d %H:%M:%S")
    except ValueError as exc:
        parser.error(f"Ungültiger Wert für --start: {exc}")

    skalierung = max(args.faelle / 2000, 0.02)

    def skaliert(basis: int, minimum: int = 1) -> int:
        return min(args.faelle * 3, max(minimum, round(basis * skalierung)))

    return GeneratorKonfiguration(
        faelle=args.faelle,
        ausgabe=args.ausgabe,
        seed=args.seed,
        startzeitpunkt=startzeitpunkt,
        fehlende_bearbeitungszeiten=skaliert(0),
        ausreisser=skaliert(5),
        platzhalter_ressource=skaliert(0),
        platzhalter_status=skaliert(0),
        fehlende_aktivitaeten=skaliert(0),
        fehlende_zeitstempel=skaliert(0),
        fehlende_lagerplaetze=skaliert(0),
        platzhalter_lagerplaetze=skaliert(0),
        identische_zeitstempel_faelle=min(args.faelle, skaliert(0)),
        ruecklaeufige_zeit_faelle=min(args.faelle, skaliert(0)),
        exakte_duplikate=skaliert(0),
    )


def verteilte_varianten(konfiguration: GeneratorKonfiguration) -> list[str]:
    anteile = [
        ("standard", konfiguration.anteil_standard),
        ("kurz", konfiguration.anteil_kurz),
        ("nacharbeit", konfiguration.anteil_nacharbeit),
        ("transport_loop", konfiguration.anteil_transport_loop),
    ]

    anzahl_je_variante: dict[str, int] = {}
    vergeben = 0
    for name, anteil in anteile[:-1]:
        anzahl = round(konfiguration.faelle * anteil)
        anzahl_je_variante[name] = anzahl
        vergeben += anzahl
    anzahl_je_variante[anteile[-1][0]] = konfiguration.faelle - vergeben

    varianten: list[str] = []
    for name, _ in anteile:
        varianten.extend([name] * anzahl_je_variante[name])

    random.shuffle(varianten)
    return varianten


def sichere_stichprobe(population: list[int] | list[str], anzahl: int) -> list[Any]:
    if not population or anzahl <= 0:
        return []
    return random.sample(population, min(anzahl, len(population)))


def dauer_fuer(aktivitaet: str) -> int:
    bereiche = {
        "Auftrag angelegt": (2, 8),
        "MOVED": (4, 28),
        "Bearbeiten": (25, 95),
        "Qualitätsprüfung": (8, 30),
        "Nacharbeit": (15, 55),
        "Verpacken": (6, 22),
        "Auftrag abgeschlossen": (1, 5),
    }
    minimum, maximum = bereiche[aktivitaet]
    return random.randint(minimum, maximum)


def ereigniszeile(
    auftrag_id: str,
    aktivitaet: str,
    zeitstempel: datetime,
    ressource: str,
    menge: int,
    status: str,
    von_lagerplatz: str | None = None,
    zu_lagerplatz: str | None = None,
) -> list[Any]:
    return [
        auftrag_id,
        aktivitaet,
        zeitstempel,
        ressource,
        dauer_fuer(aktivitaet),
        menge,
        status,
        von_lagerplatz,
        zu_lagerplatz,
    ]


def grunddaten_erzeugen(
    konfiguration: GeneratorKonfiguration,
) -> tuple[list[list[Any]], dict[str, list[int]], list[str]]:
    lagerplaetze = [
        f"C01/{gang:03d}/{fach:02d}"
        for gang in range(101, 121)
        for fach in (1, 3, 5, 7, 9, 11, 13, 15)
    ]
    pufferplaetze = [f"P01/{nummer:03d}" for nummer in range(1, 9)]
    maschinenplaetze = [f"MAS/{nummer:02d}" for nummer in range(1, 5)]
    pruefplaetze = [f"Q01/{nummer:03d}" for nummer in range(1, 3)]
    versandplaetze = [f"V01/{nummer:03d}" for nummer in range(1, 5)]

    maschinen = ["Maschine_01", "Maschine_02", "Maschine_03", "Maschine_04"]
    stapler = ["Stapler_01", "Stapler_02", "Stapler_03", "FTS_01"]
    pruefressourcen = ["Prüfplatz_01", "Prüfplatz_02"]
    verpackungsressourcen = ["Verpackung_01", "Verpackung_02"]

    varianten = verteilte_varianten(konfiguration)
    zeilen: list[list[Any]] = []
    fall_zeilen: dict[str, list[int]] = {}

    for laufnummer, variante in enumerate(varianten, start=1):
        auftrag_id = f"A{100000 + laufnummer}"
        menge = random.randint(1, 40)
        status = "abgeschlossen"
        zeitpunkt = konfiguration.startzeitpunkt + timedelta(
            minutes=(laufnummer - 1) * 17 + random.randint(0, 8)
        )

        lager = random.choice(lagerplaetze)
        maschinen_index = random.randrange(len(maschinen))
        maschine = maschinen[maschinen_index]
        maschinenplatz = maschinenplaetze[maschinen_index]
        pruef_index = random.randrange(len(pruefressourcen))
        pruefressource = pruefressourcen[pruef_index]
        pruefplatz = pruefplaetze[pruef_index]
        versand = random.choice(versandplaetze)

        folge: list[tuple[str, str, str | None, str | None]] = [
            ("Auftrag angelegt", "ERP", None, None),
        ]

        if variante == "transport_loop":
            puffer = random.choice(pufferplaetze)
            folge.extend(
                [
                    ("MOVED", random.choice(stapler), lager, puffer),
                    ("MOVED", random.choice(stapler), puffer, maschinenplatz),
                    ("Bearbeiten", maschine, None, None),
                    ("MOVED", random.choice(stapler), maschinenplatz, pruefplatz),
                    ("Qualitätsprüfung", pruefressource, None, None),
                    ("MOVED", random.choice(stapler), pruefplatz, versand),
                    ("Auftrag abgeschlossen", "ERP", None, None),
                ]
            )
        elif variante == "nacharbeit":
            folge.extend(
                [
                    ("MOVED", random.choice(stapler), lager, maschinenplatz),
                    ("Bearbeiten", maschine, None, None),
                    ("MOVED", random.choice(stapler), maschinenplatz, pruefplatz),
                    ("Qualitätsprüfung", pruefressource, None, None),
                    ("Nacharbeit", "Nacharbeitsplatz_01", None, None),
                    ("Qualitätsprüfung", pruefressource, None, None),
                    ("MOVED", random.choice(stapler), pruefplatz, versand),
                    ("Auftrag abgeschlossen", "ERP", None, None),
                ]
            )
        elif variante == "kurz":
            folge.extend(
                [
                    ("MOVED", random.choice(stapler), lager, maschinenplatz),
                    ("Bearbeiten", maschine, None, None),
                    ("MOVED", random.choice(stapler), maschinenplatz, versand),
                    ("Verpacken", random.choice(verpackungsressourcen), None, None),
                    ("Auftrag abgeschlossen", "ERP", None, None),
                ]
            )
        else:
            folge.extend(
                [
                    ("MOVED", random.choice(stapler), lager, maschinenplatz),
                    ("Bearbeiten", maschine, None, None),
                    ("MOVED", random.choice(stapler), maschinenplatz, pruefplatz),
                    ("Qualitätsprüfung", pruefressource, None, None),
                    ("MOVED", random.choice(stapler), pruefplatz, versand),
                    ("Auftrag abgeschlossen", "ERP", None, None),
                ]
            )

        zeilenindices: list[int] = []
        for aktivitaet, ressource, von, zu in folge:
            zeitpunkt += timedelta(minutes=random.randint(4, 45))
            zeilenindices.append(len(zeilen))
            zeilen.append(
                ereigniszeile(
                    auftrag_id=auftrag_id,
                    aktivitaet=aktivitaet,
                    zeitstempel=zeitpunkt,
                    ressource=ressource,
                    menge=menge,
                    status=status,
                    von_lagerplatz=von,
                    zu_lagerplatz=zu,
                )
            )
        fall_zeilen[auftrag_id] = zeilenindices

    return zeilen, fall_zeilen, varianten


def auffaelligkeiten_einbauen(
    zeilen: list[list[Any]],
    fall_zeilen: dict[str, list[int]],
    konfiguration: GeneratorKonfiguration,
) -> dict[str, int]:
    alle_indices = list(range(len(zeilen)))
    bewegungsindices = [i for i, zeile in enumerate(zeilen) if zeile[1] == "MOVED"]

    fehlende_dauer = sichere_stichprobe(alle_indices, konfiguration.fehlende_bearbeitungszeiten)
    for index in fehlende_dauer:
        zeilen[index][4] = None

    ausreisser_kandidaten = list(set(alle_indices) - set(fehlende_dauer))
    ausreisser = sichere_stichprobe(ausreisser_kandidaten, konfiguration.ausreisser)
    ausreisserwerte = [480, 720, 1440]
    for laufnummer, index in enumerate(ausreisser):
        zeilen[index][4] = ausreisserwerte[laufnummer % len(ausreisserwerte)]

    ressourcen_platzhalter = sichere_stichprobe(alle_indices, konfiguration.platzhalter_ressource)
    platzhalterfolge = ["NULL", "N/A", "-"]
    for laufnummer, index in enumerate(ressourcen_platzhalter):
        zeilen[index][3] = platzhalterfolge[laufnummer % len(platzhalterfolge)]

    status_platzhalter = sichere_stichprobe(alle_indices, konfiguration.platzhalter_status)
    for index in status_platzhalter:
        zeilen[index][6] = "-"

    fehlende_aktivitaeten = sichere_stichprobe(alle_indices, konfiguration.fehlende_aktivitaeten)
    for index in fehlende_aktivitaeten:
        zeilen[index][1] = None

    zeit_kandidaten = list(set(alle_indices) - set(fehlende_aktivitaeten))
    fehlende_zeitstempel = sichere_stichprobe(zeit_kandidaten, konfiguration.fehlende_zeitstempel)
    for index in fehlende_zeitstempel:
        zeilen[index][2] = None

    fehlende_lagerplaetze = sichere_stichprobe(
        bewegungsindices, konfiguration.fehlende_lagerplaetze
    )
    haelfte = len(fehlende_lagerplaetze) // 2
    for index in fehlende_lagerplaetze[:haelfte]:
        zeilen[index][7] = None
    for index in fehlende_lagerplaetze[haelfte:]:
        zeilen[index][8] = None

    lagerplatz_kandidaten = list(set(bewegungsindices) - set(fehlende_lagerplaetze))
    lagerplatz_platzhalter = sichere_stichprobe(
        lagerplatz_kandidaten, konfiguration.platzhalter_lagerplaetze
    )
    for laufnummer, index in enumerate(lagerplatz_platzhalter):
        if laufnummer % 3 == 0:
            zeilen[index][7] = "NULL"
        elif laufnummer % 3 == 1:
            zeilen[index][8] = "N/A"
        else:
            zeilen[index][8] = "-"

    fall_ids = list(fall_zeilen)
    gleiche_zeit_faelle = sichere_stichprobe(fall_ids, konfiguration.identische_zeitstempel_faelle)
    for fall_id in gleiche_zeit_faelle:
        indices = fall_zeilen[fall_id]
        if len(indices) >= 3:
            zeilen[indices[2]][2] = zeilen[indices[1]][2]

    rueckwaerts_kandidaten = [fall_id for fall_id in fall_ids if fall_id not in gleiche_zeit_faelle]
    ruecklaeufige_faelle = sichere_stichprobe(
        rueckwaerts_kandidaten, konfiguration.ruecklaeufige_zeit_faelle
    )
    for fall_id in ruecklaeufige_faelle:
        indices = fall_zeilen[fall_id]
        if len(indices) >= 4 and isinstance(zeilen[indices[2]][2], datetime):
            zeilen[indices[3]][2] = zeilen[indices[2]][2] - timedelta(minutes=7)

    duplikat_quellen = sichere_stichprobe(list(range(len(zeilen))), konfiguration.exakte_duplikate)
    duplikate = [list(zeilen[index]) for index in duplikat_quellen]
    zeilen.extend(duplikate)

    return {
        "fehlende_bearbeitungszeiten": len(fehlende_dauer),
        "ausreisser": len(ausreisser),
        "platzhalter_ressource": len(ressourcen_platzhalter),
        "platzhalter_status": len(status_platzhalter),
        "fehlende_aktivitaeten": len(fehlende_aktivitaeten),
        "fehlende_zeitstempel": len(fehlende_zeitstempel),
        "fehlende_lagerplaetze": len(fehlende_lagerplaetze),
        "platzhalter_lagerplaetze": len(lagerplatz_platzhalter),
        "identische_zeitstempel_faelle": len(gleiche_zeit_faelle),
        "ruecklaeufige_zeit_faelle": len(ruecklaeufige_faelle),
        "exakte_duplikate": len(duplikate),
    }


def tabellenstil_anwenden(tabelle: Table) -> None:
    tabelle.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )


def excel_erzeugen(
    konfiguration: GeneratorKonfiguration,
    zeilen: list[list[Any]],
    varianten: list[str],
    auffaelligkeiten: dict[str, int],
) -> None:
    arbeitsmappe = Workbook()
    ereignisse = arbeitsmappe.active
    assert ereignisse is not None
    ereignisse.title = "Ereignisdaten"
    ressourcen = arbeitsmappe.create_sheet("Ressourcen")
    hinweise = arbeitsmappe.create_sheet("Hinweise")

    kopf_fill = PatternFill("solid", fgColor="1F4E78")
    kopf_font = Font(bold=True, color="FFFFFF")
    untertitel_fill = PatternFill("solid", fgColor="D9EAF7")
    untertitel_font = Font(bold=True, color="1F1F1F")

    kopfzeile = [
        "Auftrag_ID",
        "Aktivität",
        "Zeitstempel",
        "Ressource",
        "Bearbeitungszeit_min",
        "Menge",
        "Status",
        "von Lagerplatz",
        "zu Lagerplatz",
    ]
    ereignisse.append(kopfzeile)
    for zeile in zeilen:
        ereignisse.append(zeile)

    for zelle in ereignisse[1]:
        zelle.fill = kopf_fill
        zelle.font = kopf_font
        zelle.alignment = Alignment(horizontal="center", vertical="center")

    ereignisse.freeze_panes = "A2"
    ereignisse.auto_filter.ref = f"A1:I{ereignisse.max_row}"

    breiten = {
        "A": 15,
        "B": 24,
        "C": 20,
        "D": 22,
        "E": 23,
        "F": 10,
        "G": 16,
        "H": 20,
        "I": 20,
    }
    for spalte, breite in breiten.items():
        ereignisse.column_dimensions[spalte].width = breite

    for zelle in ereignisse["C"][1:]:
        zelle.number_format = "yyyy-mm-dd hh:mm:ss"

    ereignisse.conditional_formatting.add(
        f"E2:E{ereignisse.max_row}",
        CellIsRule(
            operator="greaterThan",
            formula=["300"],
            fill=PatternFill("solid", fgColor="FCE4D6"),
            font=Font(color="9C0006", bold=True),
        ),
    )

    ereignistabelle = Table(
        displayName="EreignisdatenTabelle",
        ref=f"A1:I{ereignisse.max_row}",
    )
    tabellenstil_anwenden(ereignistabelle)
    ereignisse.add_table(ereignistabelle)

    ressourcenkopf = ["Ressource", "Bereich", "Kapazität", "Aktiv", "Bemerkung"]
    ressourcen.append(ressourcenkopf)
    ressourcendaten = [
        ["ERP", "Administration", 2, True, "Auftragsverwaltung"],
        ["Stapler_01", "Intralogistik", 1, True, "Flurförderzeug"],
        ["Stapler_02", "Intralogistik", 1, True, "Flurförderzeug"],
        ["Stapler_03", "Intralogistik", 1, True, "Flurförderzeug"],
        ["FTS_01", "Intralogistik", 1, True, "Fahrerloses Transportsystem"],
        ["Maschine_01", "Produktion", 1, True, "Fräsen"],
        ["Maschine_02", "Produktion", 1, True, "Drehen"],
        ["Maschine_03", "Produktion", 1, True, "Bohren"],
        ["Maschine_04", "Produktion", 1, False, "Wartung geplant"],
        ["Prüfplatz_01", "Qualität", 1, True, "Messplatz"],
        ["Prüfplatz_02", "Qualität", 1, True, "Sichtprüfung"],
        ["Nacharbeitsplatz_01", "Produktion", 1, True, "Manuelle Nacharbeit"],
        ["Verpackung_01", "Versand", 1, True, "Standardverpackung"],
        ["Verpackung_02", "Versand", 1, True, "Sonderverpackung"],
    ]
    for zeile in ressourcendaten:
        ressourcen.append(zeile)

    for zelle in ressourcen[1]:
        zelle.fill = kopf_fill
        zelle.font = kopf_font
        zelle.alignment = Alignment(horizontal="center", vertical="center")

    ressourcen.freeze_panes = "A2"
    for index, breite in enumerate((23, 18, 12, 12, 30), start=1):
        ressourcen.column_dimensions[get_column_letter(index)].width = breite

    ressourcentabelle = Table(
        displayName="RessourcenTabelle",
        ref=f"A1:E{ressourcen.max_row}",
    )
    tabellenstil_anwenden(ressourcentabelle)
    ressourcen.add_table(ressourcentabelle)

    aktivitaetszahlen = Counter(zeile[1] for zeile in zeilen if zeile[1] is not None)
    variantenzahlen = Counter(varianten)
    hinweisdaten = [
        ["Datensatz", "Konfigurierbarer ETL- und Process-Mining-Testdatensatz"],
        ["Zeilen Ereignisdaten", len(zeilen)],
        ["Fälle / Aufträge", konfiguration.faelle],
        ["Spalten Ereignisdaten", len(kopfzeile)],
        ["Startzeitpunkt", konfiguration.startzeitpunkt],
        [
            "Prozessvarianten",
            ", ".join(f"{name}: {anzahl}" for name, anzahl in sorted(variantenzahlen.items())),
        ],
        ["MOVED-Ereignisse", aktivitaetszahlen["MOVED"]],
        [
            "Transformationstest",
            "Aus 'von Lagerplatz' und 'zu Lagerplatz' kann eine neue Aktivität gebildet werden",
        ],
        ["Echte fehlende Bearbeitungszeiten", auffaelligkeiten["fehlende_bearbeitungszeiten"]],
        ["Numerische Ausreißer", auffaelligkeiten["ausreisser"]],
        ["Textuelle Platzhalter Ressource", auffaelligkeiten["platzhalter_ressource"]],
        ["Textuelle Platzhalter Status", auffaelligkeiten["platzhalter_status"]],
        ["Fehlende Aktivitäten", auffaelligkeiten["fehlende_aktivitaeten"]],
        ["Fehlende Zeitstempel", auffaelligkeiten["fehlende_zeitstempel"]],
        ["Fehlende Lagerplatzwerte", auffaelligkeiten["fehlende_lagerplaetze"]],
        ["Textuelle Lagerplatzplatzhalter", auffaelligkeiten["platzhalter_lagerplaetze"]],
        [
            "Identische Zeitstempel innerhalb eines Falls",
            auffaelligkeiten["identische_zeitstempel_faelle"],
        ],
        ["Rückläufige Zeitwerte", auffaelligkeiten["ruecklaeufige_zeit_faelle"]],
        ["Exakte Duplikate", auffaelligkeiten["exakte_duplikate"]],
        ["Zufallsseed", konfiguration.seed],
        [
            "Hinweis zur Aktivität",
            "MOVED ist bewusst generisch; Lagerplatzpaare liegen als Attribute vor",
        ],
    ]
    hinweise.append(["Testmerkmal", "Enthaltenes Beispiel / Erwartung"])
    for zeile in hinweisdaten:
        hinweise.append(zeile)

    for zelle in hinweise[1]:
        zelle.fill = kopf_fill
        zelle.font = kopf_font
        zelle.alignment = Alignment(horizontal="center", vertical="center")

    for zeile in range(2, hinweise.max_row + 1):
        hinweise.cell(zeile, 1).fill = untertitel_fill
        hinweise.cell(zeile, 1).font = untertitel_font
        hinweise.cell(zeile, 1).alignment = Alignment(vertical="top")
        hinweise.cell(zeile, 2).alignment = Alignment(wrap_text=True, vertical="top")

    hinweise.freeze_panes = "A2"
    hinweise.column_dimensions["A"].width = 40
    hinweise.column_dimensions["B"].width = 85
    hinweise["B6"].number_format = "yyyy-mm-dd hh:mm:ss"

    hinweistabelle = Table(
        displayName="HinweiseTabelle",
        ref=f"A1:B{hinweise.max_row}",
    )
    tabellenstil_anwenden(hinweistabelle)
    hinweise.add_table(hinweistabelle)

    konfiguration.ausgabe.parent.mkdir(parents=True, exist_ok=True)
    arbeitsmappe.save(konfiguration.ausgabe)


def main() -> None:
    konfiguration = argumente_lesen()
    random.seed(konfiguration.seed)

    zeilen, fall_zeilen, varianten = grunddaten_erzeugen(konfiguration)
    auffaelligkeiten = auffaelligkeiten_einbauen(
        zeilen=zeilen,
        fall_zeilen=fall_zeilen,
        konfiguration=konfiguration,
    )
    excel_erzeugen(
        konfiguration=konfiguration,
        zeilen=zeilen,
        varianten=varianten,
        auffaelligkeiten=auffaelligkeiten,
    )

    print(f"Datei erstellt: {konfiguration.ausgabe.resolve()}")
    print(f"Fälle: {konfiguration.faelle:,}".replace(",", "."))
    print(f"Ereigniszeilen inklusive Duplikaten: {len(zeilen):,}".replace(",", "."))
    print(f"Seed: {konfiguration.seed}")


if __name__ == "__main__":
    main()
