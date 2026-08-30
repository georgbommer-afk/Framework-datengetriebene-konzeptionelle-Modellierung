#!/usr/bin/env python3
"""Reproduzierbarer Generator vollständig synthetischer Produktionsrohdaten.

Der primäre Bedienweg ist der sichtbare Block ``KONFIGURATION``. Ein Aufruf mit
``python tests/Testdatagenerator.py`` erzeugt eine rohe Excel-Arbeitsmappe und
ein statisches PNML-Sollmodell relativ zur Repositorywurzel.
"""

from __future__ import annotations

import math
import random
import tempfile
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import pm4py
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pm4py.objects.petri_net.obj import Marking, PetriNet
from pm4py.objects.petri_net.utils import petri_utils

REPOSITORYWURZEL = Path(__file__).resolve().parents[1]
STANDARD_EXCEL_PFAD = REPOSITORYWURZEL / "tests/datasets/Testdatensatz_Produktion.xlsx"
STANDARD_PNML_PFAD = REPOSITORYWURZEL / "tests/datasets/Sollprozess_Produktion.pnml"
STARTZEITPUNKT = datetime(2026, 1, 5, 6, 0)
DETERMINISTISCHER_ERZEUGUNGSZEITPUNKT = datetime(2026, 8, 19, 12, 0)
PLATZHALTERWERTE = ("NULL", "N/A", "-")


@dataclass(frozen=True, slots=True)
class GeneratorKonfiguration:
    """Direkt anpassbare Parameter; sämtliche Prozentwerte liegen zwischen 0 und 100.

    Nenner: Fehlwerte und Platzhalter beziehen sich auf alle dafür zugelassenen
    Zellen der ursprünglichen Ereignisse. Ausreißer und unbekannte Ressourcen
    beziehen sich auf Aktivitätsausführungen, Duplikate auf ursprüngliche
    Ereigniszeilen, Nichtkonformität und Variantenanteile auf Produktionsaufträge.
    """

    anzahl_faelle: int
    seed: int
    fehlwerte_prozent: float
    platzhalter_prozent: float
    ausreisser_prozent: float
    duplikate_prozent: float
    unbekannte_ressourcen_prozent: float
    nichtkonforme_faelle_prozent: float
    variantenanteile: dict[str, float]
    excel_pfad: Path = STANDARD_EXCEL_PFAD
    pnml_pfad: Path = STANDARD_PNML_PFAD


# Konfigurationsblock für die lokale Ausführung in VS Code
KONFIGURATION = GeneratorKonfiguration(
    anzahl_faelle=100, #Zeilenanzahl
    seed=20260819, #das Startdatum
    fehlwerte_prozent=0.0,
    platzhalter_prozent=0.0,
    ausreisser_prozent=1.0,
    duplikate_prozent=0.0,
    unbekannte_ressourcen_prozent=0.0,
    nichtkonforme_faelle_prozent=5.0,
    variantenanteile={
        "DREH_BASIS": 10.0,
        "DREH_SCHWEISS_LACK": 18.0,
        "FRAES_BASIS": 10.0,
        "FRAES_SCHWEISS_LACK": 18.0,
        "KOMBI_SCHWEISS_LACK": 10.0,
        "KOMBI_LACK": 4.0,
        "DREH_SCHWEISS": 5.0,
        "FRAES_SCHWEISS": 5.0,
        "DREH_LACK": 5.0,
        "FRAES_LACK": 5.0,
        "KOMBI_BASIS": 2.0,
        "KOMBI_SCHWEISS": 2.0,
        "DREH_MECH_NACHARBEIT": 2.0,
        "FRAES_MECH_NACHARBEIT": 1.5,
        "KOMBI_DOPPELNACHARBEIT": 1.0,
        "DREH_LACKNACHARBEIT": 1.5,
    },
)


@dataclass(frozen=True, slots=True)
class Aktivitaetsdefinition:
    code: str
    bezeichnung: str
    fachliche_gruppe: str
    dauer_min: int
    plan_wartezeit_min: int
    ressourcen_ids: tuple[int, ...]
    status: str


@dataclass(frozen=True, slots=True)
class Ressourcendefinition:
    ressourcen_id: int
    bezeichnung: str
    ressourcentyp: str
    kapazitaet: int
    abteilung: str
    kosten_eur_je_stunde: float
    schichtmodell: str


@dataclass(frozen=True, slots=True)
class Variantendefinition:
    varianten_id: str
    beschreibung: str
    route: str
    schweissen: bool
    lackieren: bool
    mechanische_nacharbeit: bool = False
    lacknacharbeit: bool = False


@dataclass(frozen=True, slots=True)
class GeneratorErgebnis:
    """Testbares fachliches Ergebnis vor der Dateiausgabe."""

    saubere_ereignisse: tuple[dict[str, Any], ...]
    ereignisse: tuple[dict[str, Any], ...]
    ressourcen: tuple[dict[str, Any], ...]
    datenqualitaetsprotokoll: tuple[dict[str, Any], ...]
    variantenanzahlen: dict[str, int]
    auffaelligkeitsanzahlen: dict[str, int]
    nichtkonforme_faelle: tuple[int, ...]


RESSOURCEN = (
    Ressourcendefinition(
        100, "ERP-Auftragssteuerung SYN", "Informationssystem", 50, "PPS", 18.0, "24/7"
    ),
    Ressourcendefinition(
        110, "Logistikplatz SYN 1", "Arbeitsplatz", 1, "Logistik", 38.0, "2-Schicht"
    ),
    Ressourcendefinition(
        111, "Logistikplatz SYN 2", "Arbeitsplatz", 1, "Logistik", 38.0, "2-Schicht"
    ),
    Ressourcendefinition(200, "Bandsäge SYN 1", "Maschine", 1, "Mechanik", 72.0, "2-Schicht"),
    Ressourcendefinition(201, "Bandsäge SYN 2", "Maschine", 1, "Mechanik", 68.0, "2-Schicht"),
    Ressourcendefinition(210, "Drehzentrum SYN 1", "Maschine", 1, "Mechanik", 105.0, "3-Schicht"),
    Ressourcendefinition(211, "Drehzentrum SYN 2", "Maschine", 1, "Mechanik", 98.0, "2-Schicht"),
    Ressourcendefinition(220, "Fräszentrum SYN 1", "Maschine", 1, "Mechanik", 118.0, "3-Schicht"),
    Ressourcendefinition(221, "Fräszentrum SYN 2", "Maschine", 1, "Mechanik", 110.0, "2-Schicht"),
    Ressourcendefinition(230, "Bohrzentrum SYN 1", "Maschine", 1, "Mechanik", 82.0, "2-Schicht"),
    Ressourcendefinition(231, "Bohrzentrum SYN 2", "Maschine", 1, "Mechanik", 78.0, "2-Schicht"),
    Ressourcendefinition(
        240, "Entgratplatz SYN 1", "Arbeitsplatz", 1, "Mechanik", 46.0, "2-Schicht"
    ),
    Ressourcendefinition(
        241, "Entgratplatz SYN 2", "Arbeitsplatz", 1, "Mechanik", 46.0, "2-Schicht"
    ),
    Ressourcendefinition(
        250, "Messplatz Mechanik SYN 1", "Werkzeug", 1, "Qualität", 64.0, "2-Schicht"
    ),
    Ressourcendefinition(
        251, "Messplatz Mechanik SYN 2", "Werkzeug", 1, "Qualität", 64.0, "2-Schicht"
    ),
    Ressourcendefinition(
        260, "Nacharbeitsplatz Mechanik SYN", "Arbeitsplatz", 1, "Mechanik", 55.0, "2-Schicht"
    ),
    Ressourcendefinition(300, "Schweißzelle SYN 1", "Anlage", 1, "Schweißerei", 96.0, "3-Schicht"),
    Ressourcendefinition(301, "Schweißzelle SYN 2", "Anlage", 1, "Schweißerei", 90.0, "2-Schicht"),
    Ressourcendefinition(
        310, "Schleifplatz SYN 1", "Arbeitsplatz", 1, "Schweißerei", 52.0, "2-Schicht"
    ),
    Ressourcendefinition(
        311, "Schleifplatz SYN 2", "Arbeitsplatz", 1, "Schweißerei", 52.0, "2-Schicht"
    ),
    Ressourcendefinition(
        400, "Oberflächenplatz SYN", "Arbeitsplatz", 2, "Oberfläche", 49.0, "2-Schicht"
    ),
    Ressourcendefinition(410, "Lackierkabine SYN", "Anlage", 1, "Oberfläche", 125.0, "3-Schicht"),
    Ressourcendefinition(420, "Trockenkammer SYN", "Anlage", 2, "Oberfläche", 88.0, "3-Schicht"),
    Ressourcendefinition(
        430, "Prüfplatz Oberfläche SYN", "Arbeitsplatz", 1, "Qualität", 61.0, "2-Schicht"
    ),
    Ressourcendefinition(
        440, "Lacknacharbeitsplatz SYN", "Arbeitsplatz", 1, "Oberfläche", 57.0, "2-Schicht"
    ),
    Ressourcendefinition(
        500, "Vormontageplatz SYN 1", "Arbeitsplatz", 1, "Montage", 48.0, "2-Schicht"
    ),
    Ressourcendefinition(
        501, "Vormontageplatz SYN 2", "Arbeitsplatz", 1, "Montage", 48.0, "2-Schicht"
    ),
    Ressourcendefinition(
        510, "Endmontageplatz SYN 1", "Arbeitsplatz", 1, "Montage", 54.0, "2-Schicht"
    ),
    Ressourcendefinition(
        511, "Endmontageplatz SYN 2", "Arbeitsplatz", 1, "Montage", 54.0, "2-Schicht"
    ),
    Ressourcendefinition(
        520, "Funktionsprüfstand SYN", "Anlage", 1, "Qualität", 132.0, "3-Schicht"
    ),
)


AKTIVITAETEN = (
    Aktivitaetsdefinition(
        "A01", "Auftrag freigegeben", "Auftragssteuerung", 5, 0, (100,), "Pflicht"
    ),
    Aktivitaetsdefinition(
        "A02", "Material bereitstellen", "Logistik", 25, 8, (110, 111), "Pflicht"
    ),
    Aktivitaetsdefinition(
        "A03", "Zuschnitt", "Mechanische Fertigung", 45, 10, (200, 201), "Pflicht"
    ),
    Aktivitaetsdefinition(
        "A04", "Drehen", "Mechanische Fertigung", 85, 12, (210, 211), "Routenabhängig"
    ),
    Aktivitaetsdefinition(
        "A05", "Fräsen", "Mechanische Fertigung", 105, 12, (220, 221), "Routenabhängig"
    ),
    Aktivitaetsdefinition("A06", "Bohren", "Mechanische Fertigung", 55, 10, (230, 231), "Pflicht"),
    Aktivitaetsdefinition(
        "A07", "Entgraten", "Mechanische Fertigung", 28, 6, (240, 241), "Pflicht"
    ),
    Aktivitaetsdefinition(
        "A08", "Mechanische Zwischenprüfung", "Qualität", 24, 5, (250, 251), "Pflicht"
    ),
    Aktivitaetsdefinition(
        "A09", "Mechanische Nacharbeit", "Nacharbeit", 52, 12, (260,), "Optional"
    ),
    Aktivitaetsdefinition("A10", "Schweißen", "Schweißerei", 75, 15, (300, 301), "Optional"),
    Aktivitaetsdefinition(
        "A11", "Schleifen", "Schweißerei", 38, 8, (310, 311), "Nach Schweißen Pflicht"
    ),
    Aktivitaetsdefinition(
        "A12", "Oberflächenvorbereitung", "Oberfläche", 32, 12, (400,), "Optional"
    ),
    Aktivitaetsdefinition("A13", "Lackieren", "Oberfläche", 48, 10, (410,), "Optional"),
    Aktivitaetsdefinition(
        "A14", "Trocknen", "Oberfläche", 150, 5, (420,), "Nach Lackieren Pflicht"
    ),
    Aktivitaetsdefinition(
        "A15", "Oberflächenprüfung", "Qualität", 26, 6, (430,), "Nach Lackieren Pflicht"
    ),
    Aktivitaetsdefinition("A16", "Lacknacharbeit", "Nacharbeit", 42, 10, (440,), "Optional"),
    Aktivitaetsdefinition("A17", "Vormontage", "Montage", 58, 18, (500, 501), "Pflicht"),
    Aktivitaetsdefinition("A18", "Endmontage", "Montage", 92, 12, (510, 511), "Pflicht"),
    Aktivitaetsdefinition("A19", "Funktionsprüfung", "Qualität", 36, 8, (520,), "Pflicht"),
    Aktivitaetsdefinition(
        "A20", "Auftrag abgeschlossen", "Auftragssteuerung", 5, 2, (100,), "Pflicht"
    ),
)

AKTIVITAET_NACH_NAME = {wert.bezeichnung: wert for wert in AKTIVITAETEN}
RESSOURCE_NACH_ID = {wert.ressourcen_id: wert for wert in RESSOURCEN}

VARIANTEN = {
    "DREH_BASIS": Variantendefinition(
        "DREH_BASIS", "Drehteil ohne Schweiß- und Lackpfad", "DREH", False, False
    ),
    "DREH_SCHWEISS_LACK": Variantendefinition(
        "DREH_SCHWEISS_LACK", "Drehteil, geschweißt und lackiert", "DREH", True, True
    ),
    "FRAES_BASIS": Variantendefinition(
        "FRAES_BASIS", "Frästeil ohne Schweiß- und Lackpfad", "FRAES", False, False
    ),
    "FRAES_SCHWEISS_LACK": Variantendefinition(
        "FRAES_SCHWEISS_LACK", "Frästeil, geschweißt und lackiert", "FRAES", True, True
    ),
    "KOMBI_SCHWEISS_LACK": Variantendefinition(
        "KOMBI_SCHWEISS_LACK",
        "Kombinierte Dreh-/Fräsroute, geschweißt und lackiert",
        "KOMBI",
        True,
        True,
    ),
    "KOMBI_LACK": Variantendefinition(
        "KOMBI_LACK", "Kombinierte Route, nicht geschweißt, lackiert", "KOMBI", False, True
    ),
    "DREH_SCHWEISS": Variantendefinition(
        "DREH_SCHWEISS", "Drehteil, geschweißt, nicht lackiert", "DREH", True, False
    ),
    "FRAES_SCHWEISS": Variantendefinition(
        "FRAES_SCHWEISS", "Frästeil, geschweißt, nicht lackiert", "FRAES", True, False
    ),
    "DREH_LACK": Variantendefinition(
        "DREH_LACK", "Drehteil, nicht geschweißt, lackiert", "DREH", False, True
    ),
    "FRAES_LACK": Variantendefinition(
        "FRAES_LACK", "Frästeil, nicht geschweißt, lackiert", "FRAES", False, True
    ),
    "KOMBI_BASIS": Variantendefinition(
        "KOMBI_BASIS", "Kombinierte Route ohne Schweiß- und Lackpfad", "KOMBI", False, False
    ),
    "KOMBI_SCHWEISS": Variantendefinition(
        "KOMBI_SCHWEISS", "Kombinierte Route, geschweißt, nicht lackiert", "KOMBI", True, False
    ),
    "DREH_MECH_NACHARBEIT": Variantendefinition(
        "DREH_MECH_NACHARBEIT", "Drehteil mit mechanischer Nacharbeit", "DREH", True, True, True
    ),
    "FRAES_MECH_NACHARBEIT": Variantendefinition(
        "FRAES_MECH_NACHARBEIT", "Frästeil mit mechanischer Nacharbeit", "FRAES", True, True, True
    ),
    "KOMBI_DOPPELNACHARBEIT": Variantendefinition(
        "KOMBI_DOPPELNACHARBEIT",
        "Kombinierte Route mit mechanischer und Lacknacharbeit",
        "KOMBI",
        True,
        True,
        True,
        True,
    ),
    "DREH_LACKNACHARBEIT": Variantendefinition(
        "DREH_LACKNACHARBEIT", "Drehteil mit Lacknacharbeit", "DREH", False, True, False, True
    ),
}

EREIGNISSPALTEN = (
    "Quellereignis_ID",
    "Produktionsauftrag",
    "Arbeitsgang_Nr",
    "Vorgang",
    "Buchungszeitpunkt",
    "Ressourcen_ID",
    "Ist_Start",
    "Ist_Ende",
    "Soll_Start",
    "Soll_Ende",
    "Artikelnummer",
    "Produktvariante",
    "Prozessvariante",
    "Auftragsmenge",
    "Gutmenge",
    "Ausschussmenge",
    "Prioritaet",
    "Liefertermin",
    "Tatsaechlicher_Fertigstellungstermin",
    "Zwischenlagerplatz",
    "Qualitaetsstatus",
    "Nacharbeitsgrund",
    "Ruestzeit_Min",
    "Kosten_EUR",
)
FEHLWERT_ZULAESSIGE_SPALTEN = (
    "Soll_Start",
    "Soll_Ende",
    "Artikelnummer",
    "Produktvariante",
    "Prioritaet",
    "Liefertermin",
    "Tatsaechlicher_Fertigstellungstermin",
    "Zwischenlagerplatz",
    "Qualitaetsstatus",
    "Nacharbeitsgrund",
    "Ruestzeit_Min",
    "Kosten_EUR",
)
NICHTKONFORMITAETSARTEN = (
    "PFLICHTPRUEFUNG_UEBERSPRUNGEN",
    "FALSCHE_REIHENFOLGE",
    "UNZULAESSIGE_WIEDERHOLUNG",
    "VERFRUEHTER_ABSCHLUSS",
)


def _runde_anteil(grundmenge: int, prozent: float) -> int:
    """Rundet deterministisch kaufmännisch (0,5 wird aufgerundet)."""
    return math.floor(grundmenge * prozent / 100.0 + 0.5)


def validiere_konfiguration(konfiguration: GeneratorKonfiguration) -> None:
    """Validiert Grenzen, Variantennamen und die ausdrücklich geforderte 100-%-Summe."""
    if konfiguration.anzahl_faelle < 1:
        raise ValueError("anzahl_faelle muss positiv sein.")
    for name in (
        "fehlwerte_prozent",
        "platzhalter_prozent",
        "ausreisser_prozent",
        "duplikate_prozent",
        "unbekannte_ressourcen_prozent",
        "nichtkonforme_faelle_prozent",
    ):
        wert = getattr(konfiguration, name)
        if not 0 <= wert <= 100:
            raise ValueError(f"{name} muss zwischen 0 und 100 liegen.")
    unbekannt = set(konfiguration.variantenanteile) - set(VARIANTEN)
    if unbekannt:
        raise ValueError("Unbekannte Varianten: " + ", ".join(sorted(unbekannt)))
    if not konfiguration.variantenanteile:
        raise ValueError("Mindestens ein Variantenanteil muss angegeben sein.")
    if any(wert < 0 for wert in konfiguration.variantenanteile.values()):
        raise ValueError("Variantenanteile dürfen nicht negativ sein.")
    summe = math.fsum(konfiguration.variantenanteile.values())
    if not math.isclose(summe, 100.0, abs_tol=1e-9):
        raise ValueError(f"Variantenanteile müssen exakt 100 % ergeben; erhalten: {summe:g} %.")
    positive = sum(wert > 0 for wert in konfiguration.variantenanteile.values())
    if konfiguration.anzahl_faelle < positive:
        raise ValueError("Die Fallzahl muss mindestens der Zahl positiver Varianten entsprechen.")
    if konfiguration.fehlwerte_prozent + konfiguration.platzhalter_prozent > 100:
        raise ValueError("Fehlwerte und Platzhalter dürfen zusammen höchstens 100 % ergeben.")
    if len(AKTIVITAETEN) != 20 or len(AKTIVITAET_NACH_NAME) != 20:
        raise AssertionError("Der Aktivitätskatalog muss genau 20 eindeutige Aktivitäten haben.")


def _variantenanzahlen(konfiguration: GeneratorKonfiguration) -> dict[str, int]:
    """Verteilt Fälle per Largest-Remainder; jede positive Variante erscheint einmal."""
    ziele = {
        name: konfiguration.anzahl_faelle * anteil / 100.0
        for name, anteil in konfiguration.variantenanteile.items()
    }
    anzahl = {
        name: max(1, math.floor(ziel)) if konfiguration.variantenanteile[name] > 0 else 0
        for name, ziel in ziele.items()
    }
    while sum(anzahl.values()) < konfiguration.anzahl_faelle:
        name = max(anzahl, key=lambda wert: (ziele[wert] - anzahl[wert], wert))
        anzahl[name] += 1
    while sum(anzahl.values()) > konfiguration.anzahl_faelle:
        kandidaten = [name for name, wert in anzahl.items() if wert > 1]
        if not kandidaten:
            raise ValueError("Variantenanteile lassen sich für die Fallzahl nicht verteilen.")
        name = max(kandidaten, key=lambda wert: (anzahl[wert] - ziele[wert], wert))
        anzahl[name] -= 1
    return anzahl


def trace_fuer_variante(variante: Variantendefinition) -> list[str]:
    """Baut eine fachlich zulässige Route ausschließlich aus dem 20er-Katalog."""
    trace = ["Auftrag freigegeben", "Material bereitstellen", "Zuschnitt"]
    if variante.route in {"DREH", "KOMBI"}:
        trace.append("Drehen")
    if variante.route in {"FRAES", "KOMBI"}:
        trace.append("Fräsen")
    trace.extend(["Bohren", "Entgraten", "Mechanische Zwischenprüfung"])
    if variante.mechanische_nacharbeit:
        trace.extend(["Mechanische Nacharbeit", "Mechanische Zwischenprüfung"])
    if variante.schweissen:
        trace.extend(["Schweißen", "Schleifen"])
    if variante.lackieren:
        trace.extend(["Oberflächenvorbereitung", "Lackieren", "Trocknen", "Oberflächenprüfung"])
        if variante.lacknacharbeit:
            trace.extend(["Lacknacharbeit", "Lackieren", "Trocknen", "Oberflächenprüfung"])
    trace.extend(["Vormontage", "Endmontage", "Funktionsprüfung", "Auftrag abgeschlossen"])
    return trace


def _nichtkonformen_trace_erzeugen(trace: list[str], art: str) -> list[str]:
    """Erzeugt genau eine kontrollierte Abweichung ohne Fantasieaktivität."""
    ergebnis = list(trace)
    if art == "PFLICHTPRUEFUNG_UEBERSPRUNGEN":
        ergebnis.remove("Funktionsprüfung")
    elif art == "FALSCHE_REIHENFOLGE":
        links, rechts = ergebnis.index("Bohren"), ergebnis.index("Entgraten")
        ergebnis[links], ergebnis[rechts] = ergebnis[rechts], ergebnis[links]
    elif art == "UNZULAESSIGE_WIEDERHOLUNG":
        position = ergebnis.index("Endmontage")
        ergebnis.insert(position + 1, "Endmontage")
    elif art == "VERFRUEHTER_ABSCHLUSS":
        ergebnis = ergebnis[: ergebnis.index("Endmontage") + 1]
        ergebnis.append("Auftrag abgeschlossen")
    else:
        raise ValueError(f"Unbekannte Nichtkonformitätsart: {art}")
    return ergebnis


def _zwischenlagercode(vorgaenger: str | None, aktivitaet: str, wartezeit_min: float) -> str:
    if vorgaenger is None or wartezeit_min < 1:
        return "KEIN_WARTEBEREICH"
    gruppe = AKTIVITAET_NACH_NAME[aktivitaet].fachliche_gruppe
    return {
        "Logistik": "ZWL_MATERIAL",
        "Mechanische Fertigung": "ZWL_MECHANIK",
        "Qualität": "WARTEN_QUALITAET",
        "Nacharbeit": "WARTEN_NACHARBEIT",
        "Schweißerei": "ZWL_SCHWEISSEREI",
        "Oberfläche": "ZWL_OBERFLAECHE",
        "Montage": "ZWL_MONTAGE",
        "Auftragssteuerung": "KEIN_WARTEBEREICH",
    }[gruppe]


def _ressource_und_slot_waehlen(
    ressourcen_ids: tuple[int, ...],
    fallbereit: datetime,
    belegung: dict[int, list[datetime]],
    rng: random.Random,
) -> tuple[int, int, datetime]:
    """Wählt reproduzierbar den frühesten fachlich zulässigen Kapazitätsslot."""
    kandidaten: list[tuple[datetime, float, int, int]] = []
    for ressourcen_id in ressourcen_ids:
        for slot, verfuegbar in enumerate(belegung[ressourcen_id]):
            kandidaten.append((max(fallbereit, verfuegbar), rng.random(), ressourcen_id, slot))
    start, _, ressourcen_id, slot = min(kandidaten)
    return ressourcen_id, slot, start


def _artikel_fuer_route(route: str) -> tuple[str, str]:
    return {
        "DREH": ("SYN-D-100", "Synthetische Drehbaugruppe"),
        "FRAES": ("SYN-F-200", "Synthetische Fräsbaugruppe"),
        "KOMBI": ("SYN-K-300", "Synthetische Kombibaugruppe"),
    }[route]


def _saubere_ereignisse_erzeugen(
    konfiguration: GeneratorKonfiguration,
    rng: random.Random,
    variantenanzahlen: dict[str, int],
    nichtkonforme_indices: set[int],
) -> tuple[list[dict[str, Any]], dict[int, str]]:
    """Plant Fälle, Ressourcen und Zeiten zunächst ohne Zellfehler oder Ausreißer."""
    variantenfolge = [name for name, anzahl in variantenanzahlen.items() for _ in range(anzahl)]
    rng.shuffle(variantenfolge)
    belegung = {
        ressource.ressourcen_id: [STARTZEITPUNKT] * ressource.kapazitaet for ressource in RESSOURCEN
    }
    alle_zeilen: list[dict[str, Any]] = []
    nichtkonformitaet_je_fall: dict[int, str] = {}
    nc_laufnummer = 0

    for fall_index, varianten_id in enumerate(variantenfolge):
        variante = VARIANTEN[varianten_id]
        auftrag = 26000001 + fall_index
        trace = trace_fuer_variante(variante)
        if fall_index in nichtkonforme_indices:
            art = NICHTKONFORMITAETSARTEN[nc_laufnummer % len(NICHTKONFORMITAETSARTEN)]
            nc_laufnummer += 1
            trace = _nichtkonformen_trace_erzeugen(trace, art)
            nichtkonformitaet_je_fall[auftrag] = art

        menge = rng.randint(4, 24)
        ausschuss = min(
            menge - 1,
            rng.choices((0, 1, 2), weights=(0.80, 0.16, 0.04), k=1)[0],
        )
        gutmenge = menge - ausschuss
        prioritaet = rng.choices(("Normal", "Hoch", "Eilig"), weights=(0.76, 0.19, 0.05), k=1)[0]
        artikel, produkt = _artikel_fuer_route(variante.route)
        plan_bereit = STARTZEITPUNKT + timedelta(minutes=fall_index * 22 + rng.randint(0, 12))
        ist_bereit = plan_bereit + timedelta(minutes=rng.randint(0, 18))
        fallzeilen: list[dict[str, Any]] = []
        vorgaenger: str | None = None

        for schritt, aktivitaet in enumerate(trace, start=1):
            definition = AKTIVITAET_NACH_NAME[aktivitaet]
            ruestzeit = (
                0
                if aktivitaet in {"Auftrag freigegeben", "Auftrag abgeschlossen"}
                else rng.randint(3, 18)
            )
            mengenfaktor = 0.88 + min(menge, 24) * 0.012
            plan_dauer = max(2, round(definition.dauer_min * mengenfaktor + ruestzeit))
            soll_start = plan_bereit + timedelta(minutes=definition.plan_wartezeit_min)
            soll_ende = soll_start + timedelta(minutes=plan_dauer)
            fall_fruehestens = max(
                ist_bereit + timedelta(minutes=rng.randint(2, 14)),
                soll_start + timedelta(minutes=rng.randint(-8, 22)),
            )
            ressourcen_id, slot, ist_start = _ressource_und_slot_waehlen(
                definition.ressourcen_ids, fall_fruehestens, belegung, rng
            )
            ist_dauer = max(2, round(plan_dauer * rng.uniform(0.86, 1.24)))
            ist_ende = ist_start + timedelta(minutes=ist_dauer)
            belegung[ressourcen_id][slot] = ist_ende
            wartezeit = max(0.0, (ist_start - ist_bereit).total_seconds() / 60)
            ressourcenstamm = RESSOURCE_NACH_ID[ressourcen_id]
            kosten = round(
                (ist_dauer + ruestzeit) / 60 * ressourcenstamm.kosten_eur_je_stunde,
                2,
            )
            qualitaetsstatus = (
                "NACHARBEIT"
                if aktivitaet in {"Mechanische Nacharbeit", "Lacknacharbeit"}
                else "FREIGEGEBEN"
            )
            nacharbeitsgrund = {
                "Mechanische Nacharbeit": "Maßabweichung innerhalb Nacharbeitsgrenze",
                "Lacknacharbeit": "Oberflächenfehler innerhalb Nacharbeitsgrenze",
            }.get(aktivitaet, "KEINE_NACHARBEIT")
            prozessvariante = varianten_id
            if auftrag in nichtkonformitaet_je_fall:
                prozessvariante += " | NICHTKONFORM:" + nichtkonformitaet_je_fall[auftrag]
            fallzeilen.append(
                {
                    "Quellereignis_ID": f"SYN-{auftrag}-{schritt:03d}",
                    "Produktionsauftrag": auftrag,
                    "Arbeitsgang_Nr": schritt * 10,
                    "Vorgang": aktivitaet,
                    "Buchungszeitpunkt": ist_ende,
                    "Ressourcen_ID": ressourcen_id,
                    "Ist_Start": ist_start,
                    "Ist_Ende": ist_ende,
                    "Soll_Start": soll_start,
                    "Soll_Ende": soll_ende,
                    "Artikelnummer": artikel,
                    "Produktvariante": produkt,
                    "Prozessvariante": prozessvariante,
                    "Auftragsmenge": menge,
                    "Gutmenge": gutmenge,
                    "Ausschussmenge": ausschuss,
                    "Prioritaet": prioritaet,
                    "Liefertermin": None,
                    "Tatsaechlicher_Fertigstellungstermin": None,
                    "Zwischenlagerplatz": _zwischenlagercode(vorgaenger, aktivitaet, wartezeit),
                    "Qualitaetsstatus": qualitaetsstatus,
                    "Nacharbeitsgrund": nacharbeitsgrund,
                    "Ruestzeit_Min": ruestzeit,
                    "Kosten_EUR": kosten,
                }
            )
            plan_bereit = soll_ende
            ist_bereit = ist_ende
            vorgaenger = aktivitaet

        liefertermin = plan_bereit + timedelta(days=2 if prioritaet == "Eilig" else 4)
        fertigstellung = fallzeilen[-1]["Ist_Ende"]
        for zeile in fallzeilen:
            zeile["Liefertermin"] = liefertermin
            zeile["Tatsaechlicher_Fertigstellungstermin"] = fertigstellung
        alle_zeilen.extend(fallzeilen)

    return alle_zeilen, nichtkonformitaet_je_fall


def _protokolleintrag(
    fehlerart: str,
    zeile: dict[str, Any],
    spalte: str,
    original: Any,
    fehlerhaft: Any,
) -> dict[str, Any]:
    return {
        "Fehlerart": fehlerart,
        "Tabellenblatt": "Ereignisdaten",
        "Zeile_oder_Ereignisbezug": zeile["Quellereignis_ID"],
        "Spalte": spalte,
        "Urspruenglicher_Wert": original,
        "Fehlerhafter_Wert": fehlerhaft,
    }


def _zeitliche_ausreisser_einbauen(
    zeilen: list[dict[str, Any]], anzahl: int, rng: random.Random
) -> list[dict[str, Any]]:
    """Verlängert Ausführungen deutlich und verschiebt den restlichen Fall vorwärts."""
    if anzahl == 0:
        return []
    ausgewaehlt = sorted(rng.sample(range(len(zeilen)), anzahl))
    protokoll: list[dict[str, Any]] = []
    for index in ausgewaehlt:
        zeile = zeilen[index]
        original_ende = zeile["Ist_Ende"]
        dauer = original_ende - zeile["Ist_Start"]
        verschiebung = max(timedelta(minutes=180), dauer * rng.randint(5, 9))
        neues_ende = original_ende + verschiebung
        zeile["Ist_Ende"] = neues_ende
        zeile["Buchungszeitpunkt"] = neues_ende
        zeile["Kosten_EUR"] = round(float(zeile["Kosten_EUR"]) * 4.0, 2)
        auftrag = zeile["Produktionsauftrag"]
        for folge in zeilen[index + 1 :]:
            if folge["Produktionsauftrag"] != auftrag:
                if folge["Produktionsauftrag"] > auftrag:
                    break
                continue
            folge["Ist_Start"] += verschiebung
            folge["Ist_Ende"] += verschiebung
            folge["Buchungszeitpunkt"] += verschiebung
        fallzeilen = [wert for wert in zeilen if wert["Produktionsauftrag"] == auftrag]
        fertigstellung = fallzeilen[-1]["Ist_Ende"]
        for fallzeile in fallzeilen:
            fallzeile["Tatsaechlicher_Fertigstellungstermin"] = fertigstellung
        protokoll.append(
            _protokolleintrag(
                "ZEITLICHER_AUSREISSER",
                zeile,
                "Ist_Ende/Buchungszeitpunkt",
                original_ende,
                neues_ende,
            )
        )
    return protokoll


def _zellfehler_einbauen(
    zeilen: list[dict[str, Any]],
    fehlwerte: int,
    platzhalter: int,
    rng: random.Random,
) -> list[dict[str, Any]]:
    """Belegt disjunkte zulässige Zellen mit echten Leerwerten oder Platzhaltern."""
    alle_zellen = [
        (zeilenindex, spalte)
        for zeilenindex in range(len(zeilen))
        for spalte in FEHLWERT_ZULAESSIGE_SPALTEN
    ]
    ausgewaehlt = rng.sample(alle_zellen, fehlwerte + platzhalter)
    protokoll: list[dict[str, Any]] = []
    for laufnummer, (zeilenindex, spalte) in enumerate(ausgewaehlt):
        zeile = zeilen[zeilenindex]
        original = zeile[spalte]
        if laufnummer < fehlwerte:
            fehlerart, fehlerhaft = "ECHTER_FEHLWERT", None
        else:
            fehlerart = "TEXTUELLER_PLATZHALTER"
            fehlerhaft = PLATZHALTERWERTE[(laufnummer - fehlwerte) % len(PLATZHALTERWERTE)]
        zeile[spalte] = fehlerhaft
        protokoll.append(_protokolleintrag(fehlerart, zeile, spalte, original, fehlerhaft))
    return protokoll


def _unbekannte_ressourcen_einbauen(
    zeilen: list[dict[str, Any]], anzahl: int, rng: random.Random
) -> list[dict[str, Any]]:
    if anzahl == 0:
        return []
    protokoll: list[dict[str, Any]] = []
    for laufnummer, index in enumerate(sorted(rng.sample(range(len(zeilen)), anzahl)), start=1):
        zeile = zeilen[index]
        original = zeile["Ressourcen_ID"]
        fehlerhaft = 990000 + laufnummer
        zeile["Ressourcen_ID"] = fehlerhaft
        protokoll.append(
            _protokolleintrag(
                "UNBEKANNTE_RESSOURCEN_ID",
                zeile,
                "Ressourcen_ID",
                original,
                fehlerhaft,
            )
        )
    return protokoll


def _duplikate_einbauen(
    zeilen: list[dict[str, Any]], anzahl: int, rng: random.Random
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if anzahl == 0:
        return zeilen, []
    ausgewaehlt = set(rng.sample(range(len(zeilen)), anzahl))
    ergebnis: list[dict[str, Any]] = []
    protokoll: list[dict[str, Any]] = []
    for index, zeile in enumerate(zeilen):
        ergebnis.append(zeile)
        if index in ausgewaehlt:
            ergebnis.append(dict(zeile))
            protokoll.append(
                _protokolleintrag(
                    "EXAKTES_TUPEL_DUPLIKAT",
                    zeile,
                    "gesamte Ereigniszeile",
                    zeile["Quellereignis_ID"],
                    "zeilenidentische Kopie eingefügt",
                )
            )
    return ergebnis, protokoll


def _ressourcenzeilen() -> list[dict[str, Any]]:
    return [
        {
            "Ressourcen_ID": wert.ressourcen_id,
            "Ressourcenbezeichnung": wert.bezeichnung,
            "Ressourcentyp": wert.ressourcentyp,
            "Kapazitaet": wert.kapazitaet,
            "Abteilung": wert.abteilung,
            "Kosten_EUR_je_Stunde": wert.kosten_eur_je_stunde,
            "Schichtmodell": wert.schichtmodell,
        }
        for wert in RESSOURCEN
    ]


def generiere_daten(konfiguration: GeneratorKonfiguration) -> GeneratorErgebnis:
    """Erzeugt alle Daten speicherintern und ohne globalen Zufallszustand."""
    validiere_konfiguration(konfiguration)
    rng = random.Random(konfiguration.seed)
    variantenanzahlen = _variantenanzahlen(konfiguration)
    nichtkonform_anzahl = _runde_anteil(
        konfiguration.anzahl_faelle,
        konfiguration.nichtkonforme_faelle_prozent,
    )
    nichtkonforme_indices = set(rng.sample(range(konfiguration.anzahl_faelle), nichtkonform_anzahl))
    saubere_ereignisse, nichtkonformitaet_je_fall = _saubere_ereignisse_erzeugen(
        konfiguration, rng, variantenanzahlen, nichtkonforme_indices
    )
    zeilen = [dict(zeile) for zeile in saubere_ereignisse]
    urspruengliche_ereignisanzahl = len(zeilen)
    zellnenner = urspruengliche_ereignisanzahl * len(FEHLWERT_ZULAESSIGE_SPALTEN)
    sollzahlen = {
        "Echte Fehlwerte": _runde_anteil(zellnenner, konfiguration.fehlwerte_prozent),
        "Textuelle Platzhalter": _runde_anteil(zellnenner, konfiguration.platzhalter_prozent),
        "Zeitliche Ausreißer": _runde_anteil(
            urspruengliche_ereignisanzahl, konfiguration.ausreisser_prozent
        ),
        "Exakte Tupel-Duplikate": _runde_anteil(
            urspruengliche_ereignisanzahl, konfiguration.duplikate_prozent
        ),
        "Unbekannte Ressourcen-IDs": _runde_anteil(
            urspruengliche_ereignisanzahl,
            konfiguration.unbekannte_ressourcen_prozent,
        ),
        "Nicht konforme Fälle": nichtkonform_anzahl,
    }
    protokoll = _zeitliche_ausreisser_einbauen(zeilen, sollzahlen["Zeitliche Ausreißer"], rng)
    protokoll.extend(
        _zellfehler_einbauen(
            zeilen,
            sollzahlen["Echte Fehlwerte"],
            sollzahlen["Textuelle Platzhalter"],
            rng,
        )
    )
    protokoll.extend(
        _unbekannte_ressourcen_einbauen(zeilen, sollzahlen["Unbekannte Ressourcen-IDs"], rng)
    )
    for auftrag, art in sorted(nichtkonformitaet_je_fall.items()):
        bezug = next(zeile for zeile in zeilen if zeile["Produktionsauftrag"] == auftrag)
        protokoll.append(
            _protokolleintrag(
                "NICHT_KONFORMER_FALL",
                bezug,
                "Prozessfolge",
                "reguläre Variante",
                art,
            )
        )
    zeilen, duplikatprotokoll = _duplikate_einbauen(
        zeilen, sollzahlen["Exakte Tupel-Duplikate"], rng
    )
    protokoll.extend(duplikatprotokoll)
    istzahlen = {
        "Echte Fehlwerte": sum(wert["Fehlerart"] == "ECHTER_FEHLWERT" for wert in protokoll),
        "Textuelle Platzhalter": sum(
            wert["Fehlerart"] == "TEXTUELLER_PLATZHALTER" for wert in protokoll
        ),
        "Zeitliche Ausreißer": sum(
            wert["Fehlerart"] == "ZEITLICHER_AUSREISSER" for wert in protokoll
        ),
        "Exakte Tupel-Duplikate": sum(
            wert["Fehlerart"] == "EXAKTES_TUPEL_DUPLIKAT" for wert in protokoll
        ),
        "Unbekannte Ressourcen-IDs": sum(
            wert["Fehlerart"] == "UNBEKANNTE_RESSOURCEN_ID" for wert in protokoll
        ),
        "Nicht konforme Fälle": len(nichtkonformitaet_je_fall),
    }
    if istzahlen != sollzahlen:
        raise AssertionError(f"Soll-/Istanzahlen der Fehleranreicherung weichen ab: {istzahlen}")
    return GeneratorErgebnis(
        tuple(dict(zeile) for zeile in saubere_ereignisse),
        tuple(zeilen),
        tuple(_ressourcenzeilen()),
        tuple(protokoll),
        variantenanzahlen,
        istzahlen,
        tuple(sorted(nichtkonformitaet_je_fall)),
    )


def _petrinetz_erzeugen() -> tuple[PetriNet, Marking, Marking]:
    """Erzeugt das statische Sollnetz mit genau 20 sichtbaren Transitionen."""
    netz = PetriNet("Synthetischer_Sollprozess_Produktion")

    def platz(name: str) -> PetriNet.Place:
        ergebnis = PetriNet.Place(name)
        netz.places.add(ergebnis)
        return ergebnis

    def transition(name: str, label: str | None) -> PetriNet.Transition:
        ergebnis = PetriNet.Transition(name, label)
        netz.transitions.add(ergebnis)
        return ergebnis

    def verbinden(von: PetriNet.Place, trans: PetriNet.Transition, nach: PetriNet.Place) -> None:
        petri_utils.add_arc_from_to(von, trans, netz)
        petri_utils.add_arc_from_to(trans, nach, netz)

    p0 = platz("p00_start")
    p1 = platz("p01_freigegeben")
    p2 = platz("p02_material")
    p3 = platz("p03_zuschnitt")
    verbinden(p0, transition("t01", "Auftrag freigegeben"), p1)
    verbinden(p1, transition("t02", "Material bereitstellen"), p2)
    verbinden(p2, transition("t03", "Zuschnitt"), p3)

    p_dreh = platz("p04_nach_drehen")
    p_fraes = platz("p05_vor_fraesen")
    p_route = platz("p06_route_fertig")
    verbinden(p3, transition("t04", "Drehen"), p_dreh)
    verbinden(p3, transition("tau_fraesroute", None), p_fraes)
    verbinden(p_dreh, transition("tau_drehroute_ende", None), p_route)
    verbinden(p_dreh, transition("tau_kombiroute", None), p_fraes)
    verbinden(p_fraes, transition("t05", "Fräsen"), p_route)

    p_bohren = platz("p07_gebohrt")
    p_entgratet = platz("p08_entgratet")
    p_mech_entscheidung = platz("p09_mech_entscheidung")
    p_mech_fertig = platz("p10_mech_fertig")
    verbinden(p_route, transition("t06", "Bohren"), p_bohren)
    verbinden(p_bohren, transition("t07", "Entgraten"), p_entgratet)
    verbinden(
        p_entgratet,
        transition("t08", "Mechanische Zwischenprüfung"),
        p_mech_entscheidung,
    )
    verbinden(
        p_mech_entscheidung,
        transition("t09", "Mechanische Nacharbeit"),
        p_entgratet,
    )
    verbinden(
        p_mech_entscheidung,
        transition("tau_mech_freigabe", None),
        p_mech_fertig,
    )

    p_schweiss = platz("p11_geschweisst")
    p_schweiss_fertig = platz("p12_schweisspfad_fertig")
    verbinden(p_mech_fertig, transition("t10", "Schweißen"), p_schweiss)
    verbinden(p_schweiss, transition("t11", "Schleifen"), p_schweiss_fertig)
    verbinden(
        p_mech_fertig,
        transition("tau_schweissen_ueberspringen", None),
        p_schweiss_fertig,
    )

    p_lack_vorbereitet = platz("p13_lack_vorbereitet")
    p_vor_lack = platz("p14_vor_lackieren")
    p_lackiert = platz("p15_lackiert")
    p_getrocknet = platz("p16_getrocknet")
    p_lack_entscheidung = platz("p17_lack_entscheidung")
    p_oberflaeche_fertig = platz("p18_oberflaeche_fertig")
    verbinden(
        p_schweiss_fertig,
        transition("t12", "Oberflächenvorbereitung"),
        p_lack_vorbereitet,
    )
    verbinden(p_lack_vorbereitet, transition("tau_zur_lackierung", None), p_vor_lack)
    verbinden(p_vor_lack, transition("t13", "Lackieren"), p_lackiert)
    verbinden(p_lackiert, transition("t14", "Trocknen"), p_getrocknet)
    verbinden(
        p_getrocknet,
        transition("t15", "Oberflächenprüfung"),
        p_lack_entscheidung,
    )
    verbinden(
        p_lack_entscheidung,
        transition("t16", "Lacknacharbeit"),
        p_vor_lack,
    )
    verbinden(
        p_lack_entscheidung,
        transition("tau_lack_freigabe", None),
        p_oberflaeche_fertig,
    )
    verbinden(
        p_schweiss_fertig,
        transition("tau_lackieren_ueberspringen", None),
        p_oberflaeche_fertig,
    )

    p_vormontage = platz("p19_vormontage")
    p_endmontage = platz("p20_endmontage")
    p_geprueft = platz("p21_funktionsgeprueft")
    p_ende = platz("p22_ende")
    verbinden(p_oberflaeche_fertig, transition("t17", "Vormontage"), p_vormontage)
    verbinden(p_vormontage, transition("t18", "Endmontage"), p_endmontage)
    verbinden(p_endmontage, transition("t19", "Funktionsprüfung"), p_geprueft)
    verbinden(p_geprueft, transition("t20", "Auftrag abgeschlossen"), p_ende)
    return netz, Marking({p0: 1}), Marking({p_ende: 1})


def pnml_erzeugen(pfad: Path) -> None:
    """Schreibt das statische PNML über PM4Py und liest es unmittelbar testweise ein."""
    netz, anfang, ende = _petrinetz_erzeugen()
    pfad.parent.mkdir(parents=True, exist_ok=True)
    pm4py.write_pnml(netz, anfang, ende, str(pfad))
    geladenes_netz, geladener_anfang, geladenes_ende = pm4py.read_pnml(str(pfad))
    sichtbar = {wert.label for wert in geladenes_netz.transitions if wert.label is not None}
    if sichtbar != set(AKTIVITAET_NACH_NAME) or len(sichtbar) != 20:
        raise AssertionError("Das PNML enthält nicht exakt den 20er-Aktivitätskatalog.")
    if not geladener_anfang or not geladenes_ende:
        raise AssertionError("Initial- oder Finalmarkierung des PNML fehlt.")


def _tabellenstil_anwenden(tabelle: Table) -> None:
    tabelle.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )


def _blatt_fuellen(
    arbeitsmappe: Workbook,
    name: str,
    spalten: tuple[str, ...],
    zeilen: list[dict[str, Any]],
) -> None:
    blatt = arbeitsmappe.active if name == "Ereignisdaten" else arbeitsmappe.create_sheet(name)
    assert blatt is not None
    blatt.title = name
    blatt.append(list(spalten))
    for zeile in zeilen:
        blatt.append([zeile.get(spalte) for spalte in spalten])
    kopf_fill = PatternFill("solid", fgColor="004983")
    for zelle in blatt[1]:
        zelle.fill = kopf_fill
        zelle.font = Font(bold=True, color="FFFFFF")
        zelle.alignment = Alignment(horizontal="center", vertical="center")
    blatt.freeze_panes = "A2"
    blatt.auto_filter.ref = blatt.dimensions
    for spalte in blatt.columns:
        spaltenindex = spalte[0].column
        assert isinstance(spaltenindex, int)
        buchstabe = get_column_letter(spaltenindex)
        breite = min(
            55,
            max(14, max(len(str(zelle.value or "")) for zelle in spalte) + 2),
        )
        blatt.column_dimensions[buchstabe].width = breite
        for zelle in spalte[1:]:
            if isinstance(zelle.value, datetime):
                zelle.number_format = "yyyy-mm-dd hh:mm:ss"
    tabellenname = "T_" + "".join(zeichen for zeichen in name if zeichen.isalnum())
    tabelle = Table(displayName=tabellenname, ref=blatt.dimensions)
    _tabellenstil_anwenden(tabelle)
    blatt.add_table(tabelle)


def _aktivitaetsstammzeilen() -> list[dict[str, Any]]:
    zeilen = []
    for wert in AKTIVITAETEN:
        typen = sorted({RESSOURCE_NACH_ID[nummer].ressourcentyp for nummer in wert.ressourcen_ids})
        zeilen.append(
            {
                "Aktivitaetscode": wert.code,
                "Aktivitaet": wert.bezeichnung,
                "Fachliche_Gruppe": wert.fachliche_gruppe,
                "Moegliche_Ressourcentypen": ", ".join(typen),
                "Pflicht_Optionalstatus": wert.status,
                "Planbasisdauer_Min": wert.dauer_min,
            }
        )
    return zeilen


def _variantenkatalogzeilen(
    konfiguration: GeneratorKonfiguration, ergebnis: GeneratorErgebnis
) -> list[dict[str, Any]]:
    zeilen = []
    for name, zielanteil in konfiguration.variantenanteile.items():
        variante = VARIANTEN[name]
        optionen = [f"Route={variante.route}"]
        optionen.append("Schweißen" if variante.schweissen else "ohne Schweißen")
        optionen.append("Lackieren" if variante.lackieren else "ohne Lackieren")
        if variante.mechanische_nacharbeit:
            optionen.append("mechanische Nacharbeit")
        if variante.lacknacharbeit:
            optionen.append("Lacknacharbeit")
        zeilen.append(
            {
                "Varianten_ID": name,
                "Fachliche_Beschreibung": variante.beschreibung,
                "Pfade": "; ".join(optionen),
                "Zielanteil_Prozent": zielanteil,
                "Tatsaechlich_erzeugte_Anzahl": ergebnis.variantenanzahlen[name],
                "Erwartete_Konformitaet": ("Konform vor kontrollierter Fallabweichung"),
            }
        )
    for art in NICHTKONFORMITAETSARTEN:
        zeilen.append(
            {
                "Varianten_ID": "NICHTKONFORM:" + art,
                "Fachliche_Beschreibung": ("Kontrollierte Abweichung einer regulären Variante"),
                "Pfade": art,
                "Zielanteil_Prozent": konfiguration.nichtkonforme_faelle_prozent,
                "Tatsaechlich_erzeugte_Anzahl": sum(
                    art in zeile["Prozessvariante"]
                    for zeile in ergebnis.saubere_ereignisse
                    if zeile["Vorgang"] == "Auftrag freigegeben"
                ),
                "Erwartete_Konformitaet": "Nicht konform",
            }
        )
    return zeilen


def _projektrahmenzeilen() -> list[dict[str, Any]]:
    vorschlaege = {
        "Kennzeichnung": "Synthetischer Vorschlag für die manuelle Eingabe",
        "Systemtyp": "Produktion",
        "Gestalt": "Stückgut",
        "Erzeugnisstruktur": "konvergierend",
        "Materialflusskontinuität": "diskret",
        "Auftragsabwicklungsstrategie": "Make-to-Order",
        "Organisationstyp": "Werkstattfertigung mit Montagebereich",
        "Anzahl der Arbeitsgänge": "genau 20 fachliche Aktivitäten im Katalog",
        "Eingesetzte Produktionsressourcen": (
            "Maschinen, Anlagen, Arbeitsplätze, Werkzeuge und Informationssystem"
        ),
        "Untersuchungszwecke": (
            "Varianten, Durchlaufzeiten, Wartezeiten, Engpässe, Nacharbeit und Konformität"
        ),
        "Logistische Zielgrößen": ("Termintreue, Durchlaufzeit, Bestand, Auslastung und Kosten"),
        "Systemgrenze": "Von Auftragsfreigabe bis Auftragsabschluss",
        "Datenhinweis": "Ausschließlich frei erfundene synthetische Demonstrationsdaten",
    }
    return [
        {"Merkmal": name, "Vorschlag_manuelle_Eingabe": wert} for name, wert in vorschlaege.items()
    ]


def _generierungsprotokollzeilen(
    konfiguration: GeneratorKonfiguration, ergebnis: GeneratorErgebnis
) -> list[dict[str, Any]]:
    saubere = ergebnis.saubere_ereignisse
    zeilen = [
        ("Datensatzart", "Ausschließlich synthetische, frei erfundene Daten"),
        ("Seed", konfiguration.seed),
        (
            "Deterministischer Erzeugungszeitpunkt",
            DETERMINISTISCHER_ERZEUGUNGSZEITPUNKT,
        ),
        ("Fälle", konfiguration.anzahl_faelle),
        ("Ereignisse vor Fehleranreicherung", len(saubere)),
        ("Ereignisse nach Fehleranreicherung", len(ergebnis.ereignisse)),
        ("Zeitraum von", min(wert["Ist_Start"] for wert in saubere)),
        ("Zeitraum bis", max(wert["Ist_Ende"] for wert in saubere)),
        ("Variantenanteile", str(konfiguration.variantenanteile)),
        (
            "Variantenrundung",
            "Largest-Remainder; jede positive Variante mindestens einmal",
        ),
        (
            "Erwartete Join-Beziehung",
            "Ereignisdaten n:1 Ressourcenstamm über Ressourcen_ID; LEFT JOIN empfohlen",
        ),
        (
            "Case-ID-ETL",
            "Produktionsauftrag zuerst per Datentyp-Konvertierung in Text überführen",
        ),
        ("Mapping", "Produktionsauftrag -> Fall-ID"),
        ("Mapping", "Vorgang -> Aktivität"),
        ("Mapping", "Buchungszeitpunkt -> Ereigniszeitpunkt"),
        (
            "Mapping",
            "Ist_Start -> Startzeitpunkt (kanonisch start_timestamp)",
        ),
        ("Mapping", "Ist_Ende -> Endzeitpunkt (kanonisch end_timestamp)"),
        (
            "Mapping",
            "Ressourcenbezeichnung -> Ressource, erst nach LEFT JOIN",
        ),
        (
            "Zusatzattribut",
            "Soll_Start -> Ereignisattribut für Soll-Ist-Auswertung",
        ),
        (
            "Zusatzattribut",
            "Soll_Ende -> Ereignisattribut für Soll-Ist-Auswertung",
        ),
        (
            "Fehlwertzellen-Nenner",
            f"{len(saubere)} Ereignisse × {len(FEHLWERT_ZULAESSIGE_SPALTEN)} zulässige Spalten",
        ),
        (
            "Für Zellfehler zugelassene Felder",
            ", ".join(FEHLWERT_ZULAESSIGE_SPALTEN),
        ),
        (
            "Statisches Sollmodell",
            "Sollprozess_Produktion.pnml; unabhängig von Seed und Konfiguration",
        ),
    ]
    for name, wert in asdict(konfiguration).items():
        if name not in {"variantenanteile", "excel_pfad", "pnml_pfad"}:
            zeilen.append(("Konfiguration." + name, wert))
    for name, wert in ergebnis.variantenanzahlen.items():
        zeilen.append(("Variante.Istanzahl." + name, wert))
    for name, wert in ergebnis.auffaelligkeitsanzahlen.items():
        zeilen.append(("Auffälligkeit.Sollanzahl." + name, wert))
        zeilen.append(("Auffälligkeit.Istanzahl." + name, wert))
    return [{"Merkmal": name, "Wert": wert} for name, wert in zeilen]


def excel_erzeugen(
    konfiguration: GeneratorKonfiguration,
    ergebnis: GeneratorErgebnis,
    pfad: Path | None = None,
) -> Path:
    """Schreibt die bewusst rohe, mehrblättrige Excel-Arbeitsmappe."""
    ausgabe = pfad or konfiguration.excel_pfad
    ausgabe.parent.mkdir(parents=True, exist_ok=True)
    arbeitsmappe = Workbook()
    arbeitsmappe.properties.creator = "Synthetischer Produktionsdatengenerator"
    arbeitsmappe.properties.created = DETERMINISTISCHER_ERZEUGUNGSZEITPUNKT
    arbeitsmappe.properties.modified = DETERMINISTISCHER_ERZEUGUNGSZEITPUNKT
    _blatt_fuellen(
        arbeitsmappe,
        "Ereignisdaten",
        EREIGNISSPALTEN,
        list(ergebnis.ereignisse),
    )
    _blatt_fuellen(
        arbeitsmappe,
        "Ressourcenstamm",
        (
            "Ressourcen_ID",
            "Ressourcenbezeichnung",
            "Ressourcentyp",
            "Kapazitaet",
            "Abteilung",
            "Kosten_EUR_je_Stunde",
            "Schichtmodell",
        ),
        list(ergebnis.ressourcen),
    )
    _blatt_fuellen(
        arbeitsmappe,
        "Aktivitaetsstamm",
        (
            "Aktivitaetscode",
            "Aktivitaet",
            "Fachliche_Gruppe",
            "Moegliche_Ressourcentypen",
            "Pflicht_Optionalstatus",
            "Planbasisdauer_Min",
        ),
        _aktivitaetsstammzeilen(),
    )
    _blatt_fuellen(
        arbeitsmappe,
        "Variantenkatalog",
        (
            "Varianten_ID",
            "Fachliche_Beschreibung",
            "Pfade",
            "Zielanteil_Prozent",
            "Tatsaechlich_erzeugte_Anzahl",
            "Erwartete_Konformitaet",
        ),
        _variantenkatalogzeilen(konfiguration, ergebnis),
    )
    _blatt_fuellen(
        arbeitsmappe,
        "Projektrahmen",
        ("Merkmal", "Vorschlag_manuelle_Eingabe"),
        _projektrahmenzeilen(),
    )
    _blatt_fuellen(
        arbeitsmappe,
        "Generierungsprotokoll",
        ("Merkmal", "Wert"),
        _generierungsprotokollzeilen(konfiguration, ergebnis),
    )
    _blatt_fuellen(
        arbeitsmappe,
        "Datenqualitaetsprotokoll",
        (
            "Fehlerart",
            "Tabellenblatt",
            "Zeile_oder_Ereignisbezug",
            "Spalte",
            "Urspruenglicher_Wert",
            "Fehlerhafter_Wert",
        ),
        list(ergebnis.datenqualitaetsprotokoll),
    )
    arbeitsmappe.save(ausgabe)
    return ausgabe


def _konsolenausgabe(
    konfiguration: GeneratorKonfiguration,
    ergebnis: GeneratorErgebnis,
    excel_pfad: Path,
    pnml_pfad: Path,
) -> None:
    auftraege = {zeile["Produktionsauftrag"] for zeile in ergebnis.saubere_ereignisse}
    traces = {
        tuple(
            zeile["Vorgang"]
            for zeile in ergebnis.saubere_ereignisse
            if zeile["Produktionsauftrag"] == auftrag
        )
        for auftrag in auftraege
    }
    print("Synthetische Produktionsdaten erfolgreich erzeugt")
    print(f"Excel: {excel_pfad}")
    print(f"PNML:  {pnml_pfad}")
    print(f"Fälle: {konfiguration.anzahl_faelle:,}".replace(",", "."))
    ereignistext = (
        f"Ereignisse vor/nach Anreicherung: "
        f"{len(ergebnis.saubere_ereignisse):,} / {len(ergebnis.ereignisse):,}"
    )
    print(ereignistext.replace(",", "."))
    print(f"Unterschiedliche vollständige Traces: {len(traces)}")
    print(
        "Varianten: "
        + ", ".join(f"{name}={anzahl}" for name, anzahl in ergebnis.variantenanzahlen.items())
    )
    print(
        "Auffälligkeiten: "
        + ", ".join(f"{name}={anzahl}" for name, anzahl in ergebnis.auffaelligkeitsanzahlen.items())
    )


def main() -> None:
    """Führt Standardgenerierung, Dateiausgabe und PNML-Validierung aus."""
    ergebnis = generiere_daten(KONFIGURATION)
    with tempfile.TemporaryDirectory(prefix="synthetic-production-validation-") as temp:
        temporaeres_pnml = Path(temp) / "Sollprozess_Produktion.pnml"
        pnml_erzeugen(temporaeres_pnml)
        if not temporaeres_pnml.read_bytes():
            raise AssertionError("Die validierte PNML-Ausgabe ist leer.")
    excel_pfad = excel_erzeugen(KONFIGURATION, ergebnis)
    pnml_erzeugen(KONFIGURATION.pnml_pfad)
    _konsolenausgabe(
        KONFIGURATION,
        ergebnis,
        excel_pfad.resolve(),
        KONFIGURATION.pnml_pfad.resolve(),
    )


if __name__ == "__main__":
    main()
