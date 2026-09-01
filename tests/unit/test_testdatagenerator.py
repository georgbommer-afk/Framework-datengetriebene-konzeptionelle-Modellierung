"""Fachliche Unit-Tests des synthetischen Produktionsdatengenerators."""

from dataclasses import replace
from datetime import datetime
from pathlib import Path

import pm4py
import pytest
from openpyxl import load_workbook

from tests.Testdatagenerator import (
    AKTIVITAET_NACH_NAME,
    AKTIVITAETEN,
    EREIGNISSPALTEN,
    FEHLWERT_ZULAESSIGE_SPALTEN,
    KONFIGURATION,
    PLATZHALTERWERTE,
    RESSOURCE_NACH_ID,
    excel_erzeugen,
    generiere_daten,
    pnml_erzeugen,
)


@pytest.fixture(scope="module")
def standardergebnis():  # type: ignore[no-untyped-def]
    """Erzeugt den vollständigen Standardinhalt für alle lesenden Prüfungen einmal."""
    return generiere_daten(KONFIGURATION)


def _kompakte_konfiguration(**aenderungen):  # type: ignore[no-untyped-def]
    return replace(KONFIGURATION, anzahl_faelle=64, **aenderungen)


def test_seed_ist_lokal_und_reproduzierbar() -> None:
    konfiguration = _kompakte_konfiguration()
    erstes = generiere_daten(konfiguration)
    zweites = generiere_daten(konfiguration)
    anderes = generiere_daten(replace(konfiguration, seed=konfiguration.seed + 1))

    assert erstes.ereignisse == zweites.ereignisse
    assert erstes.datenqualitaetsprotokoll == zweites.datenqualitaetsprotokoll
    assert erstes.ereignisse != anderes.ereignisse


def test_standard_enthaelt_genau_den_20er_katalog_und_viele_traces(
    standardergebnis,  # type: ignore[no-untyped-def]
) -> None:
    katalog = {wert.bezeichnung for wert in AKTIVITAETEN}
    ereignisaktivitaeten = {wert["Vorgang"] for wert in standardergebnis.saubere_ereignisse}
    assert len(AKTIVITAETEN) == len(AKTIVITAET_NACH_NAME) == 20
    assert katalog == ereignisaktivitaeten
    assert "Auftrag freigegeben" in katalog
    assert "Auftrag abgeschlossen" in katalog
    assert set(EREIGNISSPALTEN) == set(standardergebnis.ereignisse[0])

    traces = {}
    for zeile in standardergebnis.saubere_ereignisse:
        traces.setdefault(zeile["Produktionsauftrag"], []).append(zeile["Vorgang"])
    assert len({tuple(wert) for wert in traces.values()}) >= 12
    for auftrag, trace in traces.items():
        if auftrag not in standardergebnis.nichtkonforme_faelle:
            assert trace[0] == "Auftrag freigegeben"
            assert trace[-1] == "Auftrag abgeschlossen"


def test_zeit_mengen_und_ressourcenlogik_des_grunddatensatzes(
    standardergebnis,  # type: ignore[no-untyped-def]
) -> None:
    je_ressource: dict[int, list[tuple[datetime, datetime]]] = {}
    positive_wartezeiten = 0
    vorheriges_ende: dict[int, datetime] = {}
    aktivitaetsressourcen: dict[str, set[int]] = {}

    for zeile in standardergebnis.saubere_ereignisse:
        assert zeile["Soll_Start"] <= zeile["Soll_Ende"]
        assert zeile["Ist_Start"] <= zeile["Ist_Ende"]
        assert zeile["Gutmenge"] + zeile["Ausschussmenge"] == zeile["Auftragsmenge"]
        assert zeile["Tatsaechlicher_Fertigstellungstermin"] >= zeile["Ist_Ende"]
        assert zeile["Kosten_EUR"] >= 0
        auftrag = zeile["Produktionsauftrag"]
        if auftrag in vorheriges_ende:
            assert zeile["Ist_Start"] >= vorheriges_ende[auftrag]
            positive_wartezeiten += zeile["Ist_Start"] > vorheriges_ende[auftrag]
        vorheriges_ende[auftrag] = zeile["Ist_Ende"]
        ressourcen_id = zeile["Ressourcen_ID"]
        assert ressourcen_id in AKTIVITAET_NACH_NAME[zeile["Vorgang"]].ressourcen_ids
        aktivitaetsressourcen.setdefault(zeile["Vorgang"], set()).add(ressourcen_id)
        je_ressource.setdefault(ressourcen_id, []).append((zeile["Ist_Start"], zeile["Ist_Ende"]))

    assert positive_wartezeiten > 0
    assert sum(len(werte) > 1 for werte in aktivitaetsressourcen.values()) >= 10
    auslastungen = [len(werte) for werte in je_ressource.values()]
    assert max(auslastungen) > min(auslastungen) * 2

    for ressourcen_id, intervalle in je_ressource.items():
        if RESSOURCE_NACH_ID[ressourcen_id].kapazitaet != 1:
            continue
        sortiert = sorted(intervalle)
        assert all(
            sortiert[index][0] >= sortiert[index - 1][1] for index in range(1, len(sortiert))
        )


def test_fehleranreicherung_ist_exakt_disjunkt_und_deaktivierbar(
    standardergebnis,  # type: ignore[no-untyped-def]
) -> None:
    protokoll = standardergebnis.datenqualitaetsprotokoll
    zellfehler = [
        wert
        for wert in protokoll
        if wert["Fehlerart"] in {"ECHTER_FEHLWERT", "TEXTUELLER_PLATZHALTER"}
    ]
    koordinaten = {(wert["Zeile_oder_Ereignisbezug"], wert["Spalte"]) for wert in zellfehler}
    assert len(koordinaten) == len(zellfehler)
    assert all(wert["Spalte"] in FEHLWERT_ZULAESSIGE_SPALTEN for wert in zellfehler)
    assert (
        sum(wert["Fehlerart"] == "ECHTER_FEHLWERT" for wert in protokoll)
        == (standardergebnis.auffaelligkeitsanzahlen["Echte Fehlwerte"])
    )
    platzhalter = [
        wert["Fehlerhafter_Wert"]
        for wert in protokoll
        if wert["Fehlerart"] == "TEXTUELLER_PLATZHALTER"
    ]
    assert set(platzhalter) <= set(PLATZHALTERWERTE)
    if len(platzhalter) >= len(PLATZHALTERWERTE):
        assert set(platzhalter) == set(PLATZHALTERWERTE)
    assert (
        len(standardergebnis.ereignisse) - len(standardergebnis.saubere_ereignisse)
        == standardergebnis.auffaelligkeitsanzahlen["Exakte Tupel-Duplikate"]
    )

    fehlerfrei = generiere_daten(
        _kompakte_konfiguration(
            fehlwerte_prozent=0,
            platzhalter_prozent=0,
            ausreisser_prozent=0,
            duplikate_prozent=0,
            unbekannte_ressourcen_prozent=0,
            nichtkonforme_faelle_prozent=0,
        )
    )
    assert not fehlerfrei.datenqualitaetsprotokoll
    assert all(wert == 0 for wert in fehlerfrei.auffaelligkeitsanzahlen.values())


def test_rohdaten_erfordern_ressourcenjoin(
    standardergebnis,  # type: ignore[no-untyped-def]
) -> None:
    assert "Ressourcen_ID" in standardergebnis.ereignisse[0]
    assert isinstance(standardergebnis.ereignisse[0]["Ressourcen_ID"], int)
    assert "Ressourcenbezeichnung" not in standardergebnis.ereignisse[0]
    ids = [wert["Ressourcen_ID"] for wert in standardergebnis.ressourcen]
    assert len(ids) == len(set(ids))


def test_excel_und_pnml_werden_vollstaendig_und_portierbar_erzeugt(
    tmp_path: Path,
    standardergebnis,  # type: ignore[no-untyped-def]
) -> None:
    excel_pfad = excel_erzeugen(KONFIGURATION, standardergebnis, tmp_path / "produktion.xlsx")
    pnml_pfad = tmp_path / "sollprozess.pnml"
    pnml_erzeugen(pnml_pfad)

    arbeitsmappe = load_workbook(excel_pfad, read_only=True, data_only=True)
    try:
        assert arbeitsmappe.sheetnames == [
            "Ereignisdaten",
            "Ressourcenstamm",
            "Aktivitaetsstamm",
            "Variantenkatalog",
            "Projektrahmen",
            "Generierungsprotokoll",
            "Datenqualitaetsprotokoll",
        ]
        assert arbeitsmappe["Ereignisdaten"].max_row == len(standardergebnis.ereignisse) + 1
        assert all(blatt.max_row > 1 for blatt in arbeitsmappe.worksheets)
    finally:
        arbeitsmappe.close()

    netz, initial, final = pm4py.read_pnml(str(pnml_pfad))
    sichtbar = [wert.label for wert in netz.transitions if wert.label is not None]
    assert len(sichtbar) == 20
    assert set(sichtbar) == {wert.bezeichnung for wert in AKTIVITAETEN}
    assert initial and final
    assert "/Users/" not in pnml_pfad.read_text(encoding="utf-8")
