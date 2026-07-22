"""Separat testbare Excel-Metadaten- und Excel-Leselogik."""

from io import BytesIO
from zipfile import BadZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from framework_mvp.domain.exceptions import Datenimportfehler
from framework_mvp.domain.models import ExcelImportparameter, TabellenblattInfo


def ermittle_tabellenblaetter(dateiinhalt: bytes) -> tuple[TabellenblattInfo, ...]:
    """Liest Blattnamen und Dimensionen ohne vollständiges DataFrame-Einlesen."""
    try:
        arbeitsmappe = load_workbook(BytesIO(dateiinhalt), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as fehler:
        raise Datenimportfehler("Die XLSX-Datei ist beschädigt oder nicht lesbar.") from fehler
    try:
        return tuple(
            TabellenblattInfo(blatt.title, blatt.max_row, blatt.max_column)
            for blatt in arbeitsmappe.worksheets
        )
    finally:
        arbeitsmappe.close()


def lese_excel(dateiinhalt: bytes, parameter: ExcelImportparameter) -> pd.DataFrame:
    """Liest genau ein ausgewähltes Tabellenblatt vollständig ein."""
    try:
        daten = pd.read_excel(
            BytesIO(dateiinhalt),
            sheet_name=parameter.tabellenblatt,
            header=parameter.kopfzeile.pandas_header,
            engine="openpyxl",
        )
    except ValueError as fehler:
        meldung = str(fehler)
        if "Worksheet" in meldung:
            raise Datenimportfehler(
                "Das ausgewählte Tabellenblatt ist nicht vorhanden."
            ) from fehler
        raise Datenimportfehler(
            "Die XLSX-Datei kann mit den gewählten Einstellungen nicht gelesen werden."
        ) from fehler
    except (BadZipFile, InvalidFileException, OSError) as fehler:
        raise Datenimportfehler("Die XLSX-Datei ist beschädigt oder nicht lesbar.") from fehler
    return daten
