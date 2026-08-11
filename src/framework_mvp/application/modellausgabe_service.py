"""Algorithmus 10: unveränderliche strukturierte Ausgabe eines validierten K*."""

import io
import re
import textwrap
from dataclasses import dataclass
from typing import Any, cast
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from framework_mvp.application.modellvalidierung_service import ModellvalidierungService
from framework_mvp.infrastructure.exceptions import Importintegritaetsfehler


@dataclass(frozen=True, slots=True)
class StrukturierteModellausgabe:
    """Reproduzierbare, nicht persistierte Dateien aus genau einem K*."""

    report_pdf: bytes | None
    report_dateiname: str | None
    excel_xlsx: bytes | None
    excel_dateiname: str | None


def _text(wert: Any) -> str:
    if wert is None:
        return "–"
    if isinstance(wert, bool):
        return "ja" if wert else "nein"
    return str(wert)


def _flach(wert: Any, pfad: str = "") -> list[tuple[str, str]]:
    """Löst verschachtelte Inhalte kontrolliert bis zu lesbaren Einzelwerten auf."""
    if isinstance(wert, dict):
        ergebnis: list[tuple[str, str]] = []
        for schluessel, inhalt in wert.items():
            teilpfad = f"{pfad}.{schluessel}" if pfad else str(schluessel)
            ergebnis.extend(_flach(inhalt, teilpfad))
        return ergebnis or [(pfad, "–")]
    if isinstance(wert, list):
        ergebnis = []
        for index, inhalt in enumerate(wert, 1):
            teilpfad = f"{pfad}[{index}]" if pfad else f"[{index}]"
            ergebnis.extend(_flach(inhalt, teilpfad))
        return ergebnis or [(pfad, "–")]
    return [(pfad, _text(wert))]


def _pdf_escape(text: str) -> bytes:
    return (
        text.encode("cp1252", errors="replace")
        .replace(b"\\", b"\\\\")
        .replace(b"(", b"\\(")
        .replace(b")", b"\\)")
    )


def _pdf_erzeugen(zeilen: list[str]) -> bytes:
    """Erzeugt einen kleinen, gültigen PDF-1.4-Report ohne weitere Abhängigkeit."""
    umgebrochen: list[str] = []
    for zeile in zeilen:
        umgebrochen.extend(
            textwrap.wrap(
                zeile,
                width=100,
                replace_whitespace=False,
                drop_whitespace=True,
                break_long_words=True,
                break_on_hyphens=False,
            )
            or [""]
        )
    seiten = [umgebrochen[index : index + 58] for index in range(0, len(umgebrochen), 58)]
    if not seiten:
        seiten = [["Validiertes konzeptionelles Modell K*"]]
    objekte: list[bytes] = []
    objekte.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    seiten_ids = [4 + index * 2 for index in range(len(seiten))]
    objekte.append(
        b"<< /Type /Pages /Count "
        + str(len(seiten)).encode()
        + b" /Kids ["
        + b" ".join(f"{wert} 0 R".encode() for wert in seiten_ids)
        + b"] >>"
    )
    objekte.append(
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>"
    )
    for seiten_id, seite in zip(seiten_ids, seiten, strict=True):
        inhalt = [b"BT", b"/F1 9 Tf", b"45 800 Td", b"12 TL"]
        for index, zeile in enumerate(seite):
            if index:
                inhalt.append(b"T*")
            inhalt.append(b"(" + _pdf_escape(zeile) + b") Tj")
        inhalt.append(b"ET")
        stream = b"\n".join(inhalt)
        objekte.append(
            (
                f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
                f"/Resources << /Font << /F1 3 0 R >> >> /Contents {seiten_id + 1} 0 R >>"
            ).encode()
        )
        objekte.append(
            b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream"
        )
    ausgabe = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for nummer, objekt in enumerate(objekte, 1):
        offsets.append(len(ausgabe))
        ausgabe.extend(f"{nummer} 0 obj\n".encode())
        ausgabe.extend(objekt)
        ausgabe.extend(b"\nendobj\n")
    xref = len(ausgabe)
    ausgabe.extend(f"xref\n0 {len(objekte) + 1}\n".encode())
    ausgabe.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        ausgabe.extend(f"{offset:010d} 00000 n \n".encode())
    ausgabe.extend(
        f"trailer\n<< /Size {len(objekte) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode()
    )
    return bytes(ausgabe)


def _report_zeilen(k_stern: dict[str, Any]) -> list[str]:
    gesamt = k_stern["gesamtvalidierung"]
    zeilen = [
        "Validiertes konzeptionelles Modell K*",
        f"Projekt: {k_stern['projekt_id']}",
        f"Modell-ID: {k_stern['k_stern_id']}",
        f"Validierungslauf: {k_stern['validierungslauf_id']}",
        f"Erstellt: {k_stern['erstellt_am']}",
        f"Validierungsstatus: {gesamt['status']}",
        f"Validierungsvermerk: {_text(gesamt.get('validierungsvermerk'))}",
        f"K-Referenz: {k_stern['k_referenz']['k_id']} / {k_stern['k_referenz']['datei_sha256']}",
        f"O-Referenz: {k_stern['o_referenz']['o_id']} / {k_stern['o_referenz']['datei_sha256']}",
        "",
    ]
    for index, bestandteil in enumerate(k_stern["modellbestandteile"], 1):
        zeilen.extend([f"{index}. {bestandteil['bezeichnung']}", "Ursprüngliche Inhalte aus K:"])
        original = bestandteil["urspruenglicher_bestandteil"]
        informationen = original.get("informationen", [])
        if not informationen:
            zeilen.append("- keine direkt übernommenen Informationen")
        for information in informationen:
            quelle = information.get("herkunftsartefakt", "–")
            referenz = information.get("strukturreferenz", "–")
            artefakt_id = information.get("herkunftsartefakt_id", "–")
            for pfad, wert in _flach(information.get("wert"), referenz):
                zeilen.append(f"- [{quelle}; {artefakt_id}] {pfad}: {wert}")
        menschlich = bestandteil.get("menschliche_eintraege", [])
        zeilen.append("Menschliche Ergänzungen/Anpassungen:")
        if not menschlich:
            zeilen.append("- keine")
        for eintrag in menschlich:
            for pfad, wert in _flach(eintrag):
                zeilen.append(f"- [menschliche Entscheidung] {pfad}: {wert}")
        zeilen.append("")
    return zeilen


def _excel_erzeugen(k_stern: dict[str, Any]) -> bytes:
    arbeitsmappe = Workbook()
    bestandteile = cast(Worksheet, arbeitsmappe.active)
    bestandteile.title = "Modellbestandteile"
    kopf = [
        "Reihenfolge",
        "Bestandteil-ID",
        "Bestandteil",
        "Inhaltstyp",
        "Strukturreferenz",
        "Herkunft oder Entscheidung",
        "Inhalt",
    ]
    bestandteile.append(kopf)
    for index, bestandteil in enumerate(k_stern["modellbestandteile"], 1):
        original = bestandteil["urspruenglicher_bestandteil"]
        informationen = original.get("informationen", [])
        if not informationen:
            bestandteile.append(
                [
                    index,
                    bestandteil["bestandteil_id"],
                    bestandteil["bezeichnung"],
                    "K",
                    "–",
                    "keine direkt übernommene Information",
                    "–",
                ]
            )
        for information in informationen:
            herkunft = (
                f"{information.get('herkunftsartefakt', '–')} · "
                f"{information.get('herkunftsartefakt_id', '–')}"
            )
            for pfad, wert in _flach(
                information.get("wert"), information.get("strukturreferenz", "")
            ):
                bestandteile.append(
                    [
                        index,
                        bestandteil["bestandteil_id"],
                        bestandteil["bezeichnung"],
                        "ursprünglich aus K",
                        pfad,
                        herkunft,
                        wert,
                    ]
                )
        for menschlicher_eintrag in bestandteil.get("menschliche_eintraege", []):
            entscheidung = menschlicher_eintrag.get(
                "entscheidung", menschlicher_eintrag.get("eintragstyp", "menschlich")
            )
            for pfad, wert in _flach(menschlicher_eintrag):
                bestandteile.append(
                    [
                        index,
                        bestandteil["bestandteil_id"],
                        bestandteil["bezeichnung"],
                        "menschliche Ergänzung/Anpassung",
                        pfad,
                        entscheidung,
                        wert,
                    ]
                )
    metadaten = arbeitsmappe.create_sheet("Metadaten_Lineage")
    metadaten.append(["Bereich", "Schlüssel", "Wert"])
    for bereich in (
        "projekt_id",
        "k_stern_id",
        "validierungslauf_id",
        "artefaktversion",
        "erstellt_am",
        "k_referenz",
        "o_referenz",
        "gesamtvalidierung",
        "eingabefingerabdruck",
        "entscheidungsfingerabdruck",
        "gesamtpruefsumme",
    ):
        for pfad, wert in _flach(k_stern.get(bereich), bereich):
            metadaten.append([bereich, pfad, wert])
    for tabellenblatt in (bestandteile, metadaten):
        for zelle in tabellenblatt[1]:
            zelle.font = Font(bold=True, color="FFFFFF")
            zelle.fill = PatternFill("solid", fgColor="35546D")
        tabellenblatt.freeze_panes = "A2"
        tabellenblatt.auto_filter.ref = tabellenblatt.dimensions
        for spaltennummer, spalte in enumerate(tabellenblatt.columns, 1):
            breite = min(max(len(_text(zelle.value)) for zelle in spalte) + 2, 60)
            tabellenblatt.column_dimensions[get_column_letter(spaltennummer)].width = breite
            for zelle in spalte:
                zelle.alignment = Alignment(vertical="top", wrap_text=True)
    puffer = io.BytesIO()
    arbeitsmappe.save(puffer)
    return puffer.getvalue()


class ModellausgabeService:
    """Erzeugt Report und/oder Excel ausschließlich aus einem erneut validierten K*."""

    def __init__(self, validierungen: ModellvalidierungService) -> None:
        self._validierungen = validierungen

    def erzeugen(
        self,
        *,
        validierungslauf_id: UUID,
        projekt_id: UUID,
        k_stern_id: UUID,
        report: bool,
        excel: bool,
    ) -> StrukturierteModellausgabe:
        if not report and not excel:
            raise Importintegritaetsfehler("Mindestens Report oder Excel muss gewählt werden.")
        k_stern = self._validierungen.uebergabe_schritt10(
            validierungslauf_id, projekt_id, k_stern_id
        )
        kurz_projekt = re.sub(r"[^A-Za-z0-9_-]", "-", str(projekt_id))[:8]
        kurz_modell = re.sub(r"[^A-Za-z0-9_-]", "-", str(k_stern_id))[:8]
        basis = f"konzeptionelles-modell-{kurz_projekt}-{kurz_modell}"
        return StrukturierteModellausgabe(
            _pdf_erzeugen(_report_zeilen(k_stern)) if report else None,
            f"{basis}.pdf" if report else None,
            _excel_erzeugen(k_stern) if excel else None,
            f"{basis}.xlsx" if excel else None,
        )
